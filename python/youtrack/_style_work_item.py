"""
basic() --- set the basic style;\n
_basic_state_style() --- set the basic state style;\n
state_styles() --- set the state styles;\n
basic_issue_style() --- set the basic cell style for issue attribute values;\n
header() --- set the header style;\n
header_text_style() --- set the header merged cell style;\n
header_no_border_style() --- set the header cross style;\n
deadline_issue_style() --- set the issue deadline style;\n
sum_style() --- set the total spent time cell style;\n
generate_from_style(name, base_style, cell_attrs, values) -> _StyleWorkItem --- set the style based on the other one;
"""

from copy import copy
from typing import Union
from openpyxl.styles.numbers import FORMAT_GENERAL, FORMAT_TEXT, FORMAT_NUMBER_00, FORMAT_DATE_XLSX14, FORMAT_NUMBER
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM
from openpyxl.styles.colors import Color, WHITE, BLACK
from openpyxl.styles.fills import PatternFill, FILL_SOLID
from openpyxl.styles.fonts import Font
from openpyxl.styles.protection import Protection
from openpyxl.cell.cell import Cell
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.workbook.workbook import Workbook


def basic():
    """
    Specify the basic style.

    :return: the basic style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "basic", True, number_format=FORMAT_GENERAL, alignment=ConstStyle.TMP_ALIGNMENT,
        border=ConstStyle.TMP_BORDER, fill=ConstStyle.TMP_FILL, font=ConstStyle.TMP_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="n"
    )


def _basic_state_style():
    """
    Specify the basic state style.

    :return: the basic style for states.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "_basic_style", False, number_format=FORMAT_NUMBER_00, alignment=ConstStyle.CENTER_ALIGNMENT,
        border=ConstStyle.THIN_BORDER, fill=ConstStyle.TMP_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="n")


def state_styles():
    """
    Specify the state styles.

    :return: the state styles.
    :rtype: list[_StyleWorkItem]
    """
    return [generate_from_style(name, _basic_state_style(), ["_is_named", "fill"], [True, fill])
            for name, fill in ConstStyle.dict_states_color.items()]


def header():
    """
    Specify the header style.

    :return: the header style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "header", True, number_format=FORMAT_TEXT, alignment=ConstStyle.CENTER_ALIGNMENT,
        border=ConstStyle.TOP_BOTTOM_BORDER, fill=ConstStyle.HEADER_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="n")


def deadline_issue_style():
    """
    Specify the issue deadline style.

    :return: the issue deadline style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "deadline_issue", False, number_format=FORMAT_DATE_XLSX14, alignment=ConstStyle.TMP_ALIGNMENT,
        border=ConstStyle.LEFT_RIGHT_TOP_BOTTOM_BORDER, fill=ConstStyle.TMP_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="d"
    )


def header_no_border_style():
    """
    Specify the header cross style.

    :return: the header cross style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "header_no_border", False, number_format=FORMAT_GENERAL, alignment=ConstStyle.NONE_ALIGNMENT,
        border=ConstStyle.TMP_BORDER, fill=ConstStyle.HEADER_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="n"
    )


def basic_issue_style():
    """
    Specify the basic cell style for issue attribute values.

    :return: the issue attribute value cell style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "basic_issue", True, number_format=FORMAT_TEXT, alignment=ConstStyle.CENTER_ALIGNMENT,
        border=ConstStyle.LEFT_RIGHT_TOP_BOTTOM_BORDER, fill=ConstStyle.TMP_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="n"
    )


def header_text_style():
    """
    Specify the header merged cell style.

    :return: the header merged cell style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "header_text", False, number_format=FORMAT_GENERAL, alignment=ConstStyle.CENTER_ALIGNMENT,
        border=ConstStyle.LEFT_TOP_BOTTOM_BORDER, fill=ConstStyle.HEADER_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="n"
    )


def sum_style():
    """
    Specify the total spent time cell style.

    :return: the total time cell style.
    :rtype: _StyleWorkItem
    """
    return _StyleWorkItem(
        "sum", False, number_format=FORMAT_NUMBER, alignment=ConstStyle.CENTER_ALIGNMENT,
        border=ConstStyle.LEFT_RIGHT_TOP_BOTTOM_BORDER, fill=ConstStyle.TMP_FILL, font=ConstStyle.THEME_FONT,
        protection=ConstStyle.TMP_PROTECTION, data_type="f"
    )


class ConstStyle:
    """
    Contain the constants for Named and Cell Styles.

    dict_states_color --- dict of states, PatternFill values
        "weekend", "deadline", "done", "active", "test", "going_start", "paused",
        "verified_closed", "going_finish", "sick", "vacation"\n
    """
    # Cell.alignment
    TMP_ALIGNMENT = Alignment(
        horizontal='left', vertical='center', wrap_text=True, shrinkToFit=None, indent=0, relativeIndent=0,
        justifyLastLine=None, readingOrder=0)
    CENTER_ALIGNMENT = Alignment(
        horizontal='center', vertical='center', wrap_text=True, shrinkToFit=None, indent=0, relativeIndent=0,
        justifyLastLine=None, readingOrder=0)
    NONE_ALIGNMENT = Alignment()
    # Cell.border
    TMP_BORDER = Border(outline=False,
                        left=Side(style=None, color=None),
                        right=Side(style=None, color=None),
                        top=Side(style=None, color=None),
                        bottom=Side(style=None, color=None),
                        diagonal=Side(style=None, color=None))
    THIN_BORDER = Border(outline=True,
                         left=Side(style=BORDER_THIN, color=Color(rgb=BLACK, type="rgb")),
                         right=Side(style=BORDER_THIN, color=Color(rgb=BLACK, type="rgb")),
                         top=Side(style=BORDER_THIN, color=Color(rgb=BLACK, type="rgb")),
                         bottom=Side(style=BORDER_THIN, color=Color(rgb=BLACK, type="rgb")),
                         diagonal=Side(style=None, color=None))
    TOP_BOTTOM_BORDER = Border(outline=True,
                               left=Side(style=None, color=None),
                               right=Side(style=None, color=None),
                               top=Side(style=BORDER_MEDIUM, color=Color(rgb=BLACK, type="rgb")),
                               bottom=Side(style=BORDER_MEDIUM, color=Color(rgb=BLACK, type="rgb")),
                               diagonal=Side(style=None, color=None))
    LEFT_RIGHT_TOP_BOTTOM_BORDER = Border(outline=True,
                                          left=Side(style=BORDER_MEDIUM, color=Color(indexed=64, type="indexed")),
                                          right=Side(style=BORDER_MEDIUM, color=Color(indexed=64, type="indexed")),
                                          top=Side(style=BORDER_THIN, color=Color(indexed=64, type="indexed")),
                                          bottom=Side(style=BORDER_THIN, color=Color(indexed=64, type="indexed")),
                                          diagonal=Side(style=None, color=None))
    LEFT_TOP_BOTTOM_BORDER = Border(outline=True,
                                    left=Side(style=BORDER_MEDIUM, color=Color(indexed=64, type="indexed")),
                                    right=Side(style=None, color=None),
                                    top=Side(style=BORDER_THIN, color=Color(indexed=64, type="indexed")),
                                    bottom=Side(style=BORDER_THIN, color=Color(indexed=64, type="indexed")),
                                    diagonal=Side(style=None, color=None))
    # Cell.fill
    TMP_FILL = PatternFill(fill_type=None,
                           fgColor=Color(rgb=WHITE, type='rgb'),
                           bgColor=Color(rgb=WHITE, type='rgb'))
    HEADER_FILL = PatternFill(fill_type=FILL_SOLID,
                              fgColor=Color(theme=3, tint=0.5999938962981048, type='theme'),
                              bgColor=Color(indexed=64, type='indexed'))
    # Cell.font
    TMP_FONT = Font(name='Times New Roman', charset=204, family=1, color=Color(rgb=BLACK, type='rgb'), size=11)
    THEME_FONT = Font(name='Times New Roman', charset=204, family=1, color=Color(theme=1, type='theme'), size=11)
    # Cell.protection
    TMP_PROTECTION = Protection(locked=False, hidden=False)
    # states and PatternFill values
    dict_states_color: dict[str, PatternFill] = {
        "weekend": PatternFill(fill_type=FILL_SOLID,
                               fgColor=Color(rgb="FFFD5635", type='rgb'),
                               bgColor=Color(rgb=WHITE, type='rgb')),
        "deadline": PatternFill(fill_type=FILL_SOLID,
                                fgColor=Color(rgb="FFFF0D0D", type='rgb'),
                                bgColor=Color(rgb=WHITE, type='rgb')),
        "done": PatternFill(fill_type=FILL_SOLID,
                            fgColor=Color(theme=9, tint=0.0, type='theme'),
                            bgColor=Color(indexed=64, type='indexed')),
        "active": PatternFill(fill_type=FILL_SOLID,
                              fgColor=Color(theme=9, tint=0.3999755851924192, type='theme'),
                              bgColor=Color(indexed=64, type='indexed')),
        "test": PatternFill(fill_type=FILL_SOLID,
                            fgColor=Color(rgb="FF117D68", type='rgb'),
                            bgColor=Color(rgb=WHITE, type='rgb')),
        "going_start": PatternFill(fill_type=FILL_SOLID,
                                   fgColor=Color(rgb="FF00B0F0", type='rgb'),
                                   bgColor=Color(rgb=WHITE, type='rgb')),
        "paused": PatternFill(fill_type=FILL_SOLID,
                              fgColor=Color(theme=0, tint=-0.499984740745262, type='theme'),
                              bgColor=Color(indexed=64, type='indexed')),
        "verified_closed": PatternFill(fill_type=FILL_SOLID,
                                       fgColor=Color(theme=1, tint=0.249977111117893, type='theme'),
                                       bgColor=Color(indexed=64, type='indexed')),
        "going_finish": PatternFill(fill_type=FILL_SOLID,
                                    fgColor=Color(theme=4, tint=-0.249977111117893, type='theme'),
                                    bgColor=Color(indexed=64, type='indexed')),
        "sick": PatternFill(fill_type=FILL_SOLID,
                            fgColor=Color(theme=7, tint=-0.249977111117893, type='theme'),
                            bgColor=Color(indexed=64, type='indexed')),
        "vacation": PatternFill(fill_type=FILL_SOLID,
                                fgColor=Color(rgb="FFFFFF00", type='rgb'),
                                bgColor=Color(rgb=WHITE, type='rgb'))
    }


class _StyleWorkItem:
    """
    Define the style of the cell.

    Constants:
        list_attrs --- the list of attributes:
            "name", "_is_named", "number_format", "alignment", "border", "fill", "font", "protection", "data_type"

    Params:
        name --- the cell style name, name, str;\n
        _is_named --- the cell named style flag, -, str;\n
        number_format --- the cell number format, fmt;\n
        alignment --- the cell alignment, Alignment;\n
        border --- the cell border, Border;\n
        fill --- the cell fill in, Fill;\n
        font --- the cell font, Font;\n
        protection --- the cell protection, Protection;\n
        data_type --- the cell data type, data_type, str;\n

    Properties:
        get_named --- convert to the NamedStyle instance;\n

    Functions:
        set_style(cell/coord) --- set the style of the cell;\n
    """

    list_attrs: tuple[str] = (
        "name", "_is_named", "number_format", "alignment", "border", "fill", "font", "protection", "data_type"
    )

    def __init__(
            self,
            name: str = None,
            _is_named: bool = False, *,
            number_format: str = None,
            alignment: Alignment = None,
            border: Border = None,
            fill: PatternFill = None,
            font: Font = None,
            protection: Protection = None,
            data_type: str = None):

        if number_format is None:
            number_format = FORMAT_GENERAL
        if alignment is None:
            alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if border is None:
            border = Border(outline=False, left=Side(), right=Side(), top=Side(), bottom=Side(), diagonal=Side())
        if fill is None:
            fill = PatternFill(fill_type=FILL_SOLID, fgColor=Color(rgb=WHITE, type='rgb'),
                               bgColor=Color(rgb=WHITE, type='rgb'))
        if font is None:
            font = Font(name='Calibri', charset=204, family=2, color=Color(rgb=BLACK, type='rgb'), size=11)
        if protection is None:
            protection = Protection(locked=False, hidden=False)
        if data_type is None:
            data_type = "n"

        self.name = name
        self._is_named = _is_named
        self.number_format = number_format
        self.alignment = alignment
        self.border = border
        self.fill = fill
        self.font = font
        self.protection = protection
        self.data_type = data_type

    def __str__(self):
        return f"name = {self.name}, number_format = {self.number_format}, alignment = {self.alignment}, " \
               f"border = {self.border}, fill = {self.fill}, font = {self.font}, protection = {self.protection}, " \
               f"data_type = {self.data_type}"

    def __repr__(self):
        return f"StyleWorkItem(name={self.name}, number_format={self.number_format}, alignment={self.alignment}, " \
               f"border={self.border}, fill={self.fill}, font={self.font}, protection={self.protection}), " \
               f"data_type={self.data_type}"

    def __hash__(self):
        return hash((self.name, self.number_format, self.alignment, self.border, self.fill,
                     self.font, self.protection, self.data_type))

    def __eq__(self, other):
        if isinstance(other, _StyleWorkItem):
            return self.name == other.name
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, _StyleWorkItem):
            return self.name != other.name
        else:
            return NotImplemented

    def __bool__(self):
        return any((self.name, self.number_format, self.alignment, self.border, self.fill,
                    self.font, self.protection, self.data_type) is not None)

    @property
    def get_named(self):
        """
        Convert the _StyleWorkItem instance to the NamedStyle one.

        :return: the NamedStyle instance or self.
        :rtype: NamedStyle or _StyleWorkItem
        """
        if self._is_named:
            return NamedStyle(
                name=self.name, font=self.font, fill=self.fill, border=self.border, alignment=self.alignment,
                number_format=self.number_format, protection=self.protection)
        else:
            return self

    def set_style(self, cell: Cell):
        """
        Specify the style of the cell.

        :param cell: the cell or the cell coordinates
        :type cell: Cell
        :return: None.
        """
        # apply the alignment
        cell.alignment = copy(self.alignment)
        # apply the border
        cell.border = copy(self.border)
        # apply the fill
        cell.fill = copy(self.fill)
        # apply the font
        cell.font = copy(self.font)
        # apply the protection
        cell.protection = copy(self.protection)
        # apply the number format
        cell.number_format = copy(self.number_format)


def generate_from_style(name: str, base_style: _StyleWorkItem, attrs: list = None, values: list = None):
    """
    Modify the basic style.

    :param str name: the style name
    :param _StyleWorkItem base_style: the base of the style
    :param list attrs: the attributes to change
    :param list values: the attribute values
    :return: the style.
    :rtype: _StyleWorkItem
    """
    check_attrs: bool = (attrs is not None and all(attr in _StyleWorkItem.list_attrs for attr in attrs))
    check_values: bool = values is not None
    check_length: bool = len(attrs) == len(values)
    if not (check_attrs and check_values and check_length):
        return basic()
    else:
        style = copy(base_style)
        style.name = name
        for attr, value in zip(attrs, values):
            setattr(style, attr, value)
        return style


class _StyleWorkItemList:
    """
    Define the style list.

    Params:
        name --- the list name;\n
        styles --- the dictionary of the styles,
            dict[style_name, _StyleWorkItem];\n

    Functions:
        set_style(style_name, cell) --- set the style to the cell;\n
    """

    def __init__(self, name: str):
        self.name = name
        self.styles: dict[str, Union[_StyleWorkItem, NamedStyle]] = dict()
        self.styles["basic"] = basic().get_named
        self.styles["_basic_style"] = _basic_state_style().get_named
        self.styles["basic_issue"] = basic_issue_style().get_named
        self.styles["header"] = header().get_named
        self.styles["header_no_border"] = header_no_border_style().get_named
        self.styles["header_text"] = header_text_style().get_named
        self.styles["deadline_issue"] = deadline_issue_style().get_named
        self.styles["sum"] = sum_style().get_named
        style: Union[_StyleWorkItem, NamedStyle]
        for style in state_styles():
            self.styles[style.name] = style.get_named

    def __str__(self):
        str_styles = [style_name for style_name in self.styles.keys()]
        unified_str_styles = ", ".join(str_styles)
        return f"_StyleWorkItemList: name = {self.name}, styles: {unified_str_styles}"

    def __repr__(self):
        repr_styles = [repr(style_name) for style_name in self.styles.keys()]
        unified_repr_styles = ",".join(repr_styles)
        return f"_StyleWorkItemList(name={self.name}, styles=[{unified_repr_styles}])"

    def __hash__(self):
        return hash(self.name)

    def __eq__(self, other):
        if isinstance(other, _StyleWorkItemList):
            return self.name == other.name
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, _StyleWorkItemList):
            return self.name != other.name
        else:
            return NotImplemented

    def set_style(self, style_name: str, cell: Cell):
        """
        Specify the style of the cell.

        :param str style_name: the style name
        :param cell: the cell
        :type cell: Cell
        :return: None.
        """
        style = self.styles[style_name]
        if isinstance(style, NamedStyle):
            cell.style = style.name
        elif isinstance(style, _StyleWorkItem):
            style.set_style(cell)
        else:
            print(f"Something went wrong. Cell {cell.coordinate}. Style: {style_name}.")
