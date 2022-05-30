from typing import Any
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from yt_requests import User, Issue, IssueWorkItem
from xl_table import ExcelProp, TableCell


def convert_issue_state(state: str) -> str:
    """
    Converts the state to the table headers.

    :param str state: the issue state
    :return: the modified state.
    :rtype: str
    """
    dict_states = {
        "Active": "Active",
        "New": "New/Paused",
        "Paused": "New/Paused",
        "Canceled": "New/Paused",
        "Discuss": "New/Paused",
        "Done": "Done/Test",
        "Test": "Done/Test",
        "Review": "Done/Test",
        "Verified": "Verified",
        "Closed": "Verified"
    }
    if state in dict_states.keys():
        return dict_states[state]
    else:
        print(f"Unspecified state {state} is found.")
        return state


class JoinUserXL:
    """
    Class params:
        attrs_column --- the attributes to get;\n
        dict_attr_column --- the dictionary of the issue attributes and columns;\n

    Params:
        user --- the User instance;\n
        excel_prop --- the ExcelProp instance;\n

    Properties:
        ws --- the worksheet;\n
        issue_names --- the issue names;\n
        work_items --- the issue work items;\n
        table_names --- the table cell names;\n

    Functions:
        __index(state) --- get the state index in the headers;\n
        new_row(state) --- get the new issue row in the table;\n
        add_row(row) --- add the row to the table;\n
        issue_by_name(issue_name) --- get the Issue instance by its name if exists;\n
        add_issue(issue_name) --- add the issue to the table;\n
        add_all_issues() --- add all issues to the table;\n
        _parse_attr_seq(attrs, values, row) --- add the attribute values to the table;\n
        modify_issue(issue_name) --- modify the attribute values in the table;\n
        modify_all_issues() --- modify all attribute values in the table;\n
        __get_issue_by_name(issue_name) --- get the Issue and TableCell instances by the issue name;\n
        modify_state(issue_name) --- modify the issue state;\n
        work_items_by_name(issue_name) --- get the IssueWorkItem instances by the issue name;\n
        add_work_items_issue(issue_name) --- add the work items for the issue;\n
        add_new_work_items() --- add the new work items;\n
        set_work_item_styles() --- set the styles to the work items;\n
    """
    dict_attr_column = {"parent": "B", "issue": "C", "summary": "D", "deadline": "E", "commentary": "NT"}
    attrs_column = ("parent", "summary", "deadline")

    def __init__(self, user: User, excel_prop: ExcelProp):
        self.user = user
        self.excel_prop = excel_prop

    def __str__(self):
        return f"User: {self.user.login}, Excel table: {self.excel_prop.name}"

    def __repr__(self):
        return f"JoinUserXL({repr(self.user)}, {repr(self.excel_prop)})"

    @property
    def ws(self) -> Worksheet:
        """
        Get the worksheet.

        :return: the worksheet.
        :rtype: Worksheet
        """
        return self.excel_prop.ws

    def __index(self, state: str) -> int:
        """
        Get the state index.

        :param str state: the state
        :return: the header index.
        :rtype: int
        """
        return self.excel_prop.dict_headers_short[state]

    def new_row(self, state: str) -> int:
        """
        Get the row for the new issue in the table.

        :param str state: the issue state
        :return: the row number.
        :rtype: int
        """
        return self.excel_prop.headers_row[self.__index(convert_issue_state(state)) + 1] - 1

    def add_row(self, row: int):
        """
        Add the row to the table.

        :param int row: the row number
        :return: None.
        """
        self.ws.insert_rows(row)
        formulae = f"=SUM(G{row}:NR{row})"
        self.ws[f"NS{row}"].value = formulae

    @property
    def issue_names(self) -> set[str]:
        """
        Get the issue names.

        :return: the issue names.
        :rtype: set[str]
        """
        return set(self.user.dict_issue.keys())

    def issue_by_name(self, issue_name: str) -> Issue:
        """
        Get the Issue instance by its name if exists.

        :param str issue_name: the issue name
        :return: the issue instance if exists.
        :rtype: Issue or None
        """
        return self.user.dict_issue[issue_name]

    @property
    def work_items(self) -> list[IssueWorkItem]:
        """
        Get the IssueWorkItem instances.

        :return: the issue work items.
        :rtype: list[IssueWorkItem]
        """
        return list(self.user.dict_issue_work_item.values())

    @property
    def table_names(self) -> set[str]:
        """
        Get the table cell names.

        :return: the table cell names.
        :rtype: set[str]
        """
        return set(self.excel_prop.dict_table_cell.keys())

    def table_cell_by_name(self, issue_name: str) -> TableCell:
        """
        Get the TableCell instance by its name if exists.

        :param str issue_name: the table cell name
        :return: the table cell instance if exists.
        :rtype: TableCell or None
        """
        return self.excel_prop.dict_table_cell[issue_name]

    def add_issue(self, issue_name: str):
        """
        Add the issue to the table.\n

        issue, state, summary, parent, deadline

        :param str issue_name: the YouTrack issue name
        :return: None.
        """
        # prepare the new row
        issue = self.issue_by_name(issue_name)
        add_row = self.new_row(issue.state)
        self.add_row(add_row)
        # add the values to the associated cells
        attrs = ["parent", "issue", "summary", "deadline"]
        values = [getattr(issue, attr) for attr in attrs]
        self._parse_attr_seq(attrs, values, add_row)

    def add_all_issues(self):
        """Add all new issues."""
        names_add = self.issue_names.difference(self.table_names)
        for issue_name in names_add:
            self.add_issue(issue_name)

    def _parse_attr_seq(self, attrs: list[str], values: list[Any], row: int):
        """
        Add the attribute values to the table.

        :param attrs: the attribute names
        :type attrs: list[str]
        :param values: the attribute values
        :type values: list[Any]
        :param int row: the row in the table
        :return: None.
        """
        if len(attrs) != len(values):
            print("Improper lengths.")
            return
        for attr, value in zip(attrs, values):
            if attr not in self.dict_attr_column.keys():
                continue
            else:
                if value is not None:
                    column_letter = self.dict_attr_column[attr]
                    self.ws[f"{column_letter}{row}"].value = value
                else:
                    print("The None value is not assigned.")

    def modify_issue(self, issue_name: str):
        """
        Modify the attribute values in the table.

        :param str issue_name: the YouTrack issue name
        :return: None.
        """
        issue, table_cell = self.__get_issue_by_name(issue_name)
        row = table_cell.row
        modify_attrs = [attr for attr in self.attrs_column if getattr(table_cell, attr) != getattr(issue, attr)]
        modify_values = [getattr(issue, attr) for attr in modify_attrs]
        self._parse_attr_seq(modify_attrs, modify_values, row)

    def modify_all_issues(self):
        """Modify all issues in the table."""
        for issue_name in self.issue_names:
            self.modify_issue(issue_name)
            self.modify_state(issue_name)

    def __get_issue_by_name(self, issue_name: str) -> tuple[Issue, TableCell]:
        """
        Get the Issue and TableCell instances by the issue name.

        :param str issue_name: the issue name
        :return: the Issue and TableCell instances.
        :rtype: tuple[Issue, TableCell]
        """
        return self.issue_by_name(issue_name), self.table_cell_by_name(issue_name)

    def modify_state(self, issue_name: str):
        """
        Modify the issue state.

        :param str issue_name: the issue name
        :return: None.
        """
        issue, table_cell = self.__get_issue_by_name(issue_name)
        if convert_issue_state(issue.state) != table_cell.state:
            row = table_cell.row
            add_row = self.new_row(issue.state)
            self.add_row(add_row)
            for cell in self.excel_prop.cell_in_range(f"B{row}", f"NR{row}"):
                self.excel_prop.replace_cell(from_=cell, to_=self.ws[f"{cell.column_letter}{add_row}"])
            self.excel_prop.replace_cell(from_=f"NT{row}", to_=f"NT{add_row}")
            self.ws.delete_rows(row)
        else:
            print(f"The issue {issue_name} has not changed the state.")

    def work_items_by_name(self, issue_name: str) -> list[IssueWorkItem]:
        """
        Get the IssueWorkItem instances by the issue name.

        :param str issue_name: the issue name
        :return: the issue work items.
        :rtype: list[IssueWorkItem]
        """
        return self.user.dict_issue_work_item[issue_name]

    def add_work_items_issue(self, issue_name: str):
        """
        Add the work items for the issue.

        :param str issue_name: the issue name
        :return: None.
        """
        table_cell = self.table_cell_by_name(issue_name)
        work_item_dates = set(work_item.date for work_item in self.work_items_by_name(issue_name))
        table_cell_dates = set(date for _, date, _ in table_cell.work_items())
        new_work_items_dates = work_item_dates.difference(table_cell_dates)
        new_work_items = [work_item.to_tuple() for work_item in self.work_items
                          if work_item.date in new_work_items_dates]
        for work_item in new_work_items:
            table_cell.add_work_item(work_item)

    def add_new_work_items(self):
        """Add the new work items."""
        work_items_names = [work_item.issue for work_item in self.work_items]
        for issue_name in work_items_names:
            self.add_work_items_issue(issue_name)

    def set_work_item_styles(self):
        """Set the styles to the work items."""
        for issue_name in self.issue_names:
            table_cell = self.table_cell_by_name(issue_name)
            for work_item in table_cell.work_items():
                cell: Cell = table_cell.mapping_work_item_cell(work_item)
                style_name = table_cell.proper_work_item_style(work_item)
                table_cell.set_cell_style(style_name, cell)


def main():
    pass


if __name__ == "__main__":
    main()
