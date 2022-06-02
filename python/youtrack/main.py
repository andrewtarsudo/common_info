import datetime
from typing import Any, Optional
import openpyxl
from pathlib import Path
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from yt_config import UserConfig
from _style_work_item import _StyleWorkItemList
from yt_requests import User, Issue, IssueWorkItem
from xl_table import ExcelProp, TableCell


def convert_issue_state(state: str) -> str:
    """
    Converts the state to the table headers.

    :param str state: the issue state
    :return: the modified state.
    :rtype: str
    """
    dict_states = {
        "Active": "Active",
        "New": "New/Paused",
        "Paused": "New/Paused",
        "Canceled": "New/Paused",
        "Discuss": "New/Paused",
        "Done": "Done/Test",
        "Test": "Done/Test",
        "Review": "Done/Test",
        "Verified": "Verified",
        "Closed": "Verified"
    }
    if state in dict_states.keys():
        return dict_states[state]
    else:
        print(f"Unspecified state {state} is found.")
        return state


class JoinUserXL:
    """
    Join the User and ExcelProp instances.

    Class params:
        dict_attr_column --- the dictionary of the issue attributes and columns;\n
        attrs_column --- the attributes to get;\n

    Params:
        user --- the User instance;\n
        excel_prop --- the ExcelProp instance;\n

    Properties:
        ws --- the worksheet;\n
        issue_names --- the issue names;\n
        work_item_names --- the issue work item names;\n
        issues_to_add --- the issue names to add to the table;\n
        issues_to_update_state --- the issue names to update the state;\n
        issues_to_modify --- the issue names to modify;\n

    Functions:
        __index(state) --- get the state index in the headers;\n
        new_row(state) --- get the new issue row in the table;\n
        add_all_issues() --- add all issues to the table;\n
        _parse_attr_seq(attrs, values, row) --- add the attribute values to the table;\n
        modify_all_issues() --- modify all attribute values in the table;\n
        modify_table_cell(issue_name, state) --- modify the table cell state;\n
        table_cell_item(issue_name) --- get the TableCell instance by the issue name;\n
        __get_issue_by_name(issue_name) --- get the Issue and TableCell instances by the issue name;\n
        add_new_work_items() --- add the new work items;\n
        set_work_item_styles() --- set the styles to the work items;\n
        update_states() --- update the issue states in the table;\n
    """
    dict_attr_column = {"parent": "B", "issue": "C", "summary": "D", "deadline": "E", "commentary": "NT"}
    attrs_column = ("parent", "summary", "deadline")

    def __init__(self, user: User, excel_prop: ExcelProp):
        self.user = user
        self.excel_prop = excel_prop

    def __str__(self):
        return f"User: {self.user.login}, Excel table: {self.excel_prop.name}"

    def __repr__(self):
        return f"JoinUserXL({repr(self.user)}, {repr(self.excel_prop)})"

    @property
    def ws(self) -> Worksheet:
        """
        Get the worksheet.

        :return: the worksheet.
        :rtype: Worksheet
        """
        return self.excel_prop.ws

    def __index(self, state: str) -> int:
        """
        Get the state index.

        :param str state: the state
        :return: the header index.
        :rtype: int
        """
        return self.excel_prop.dict_headers_short[state]

    def new_row(self, state: str) -> int:
        """
        Get the row for the new issue in the table.

        :param str state: the issue state
        :return: the row number.
        :rtype: int
        """
        return self.excel_prop.headers_row[self.__index(convert_issue_state(state)) + 1] - 1

    @property
    def work_item_names(self) -> list[str]:
        """
        Get the IssueWorkItem instance names.

        :return: the issue work item names.
        :rtype: list[str]
        """
        return list(self.user.dict_issue_work_item.keys())

    @property
    def issues_to_add(self) -> set[str]:
        """
        Specify the issue names to add to the table.

        :return: the issue names.
        :rtype: set[str]
        """
        return self.user.issue_names().difference(self.excel_prop.table_cell_names())

    @property
    def issues_to_update_state(self) -> set[str]:
        """
        Specify the issue names to update the state.

        :return: the issue names.
        :rtype: set[str]
        """
        return self.excel_prop.table_cell_names().difference(self.user.issue_names())

    @property
    def issues_to_modify(self) -> set[str]:
        """
        Specify the issue names to modify.

        :return: the issue names.
        :rtype: set[str]
        """
        return set.intersection(self.user.issue_names(), self.excel_prop.table_cell_names())

    def add_all_issues(self):
        """Add all new issues."""
        for issue_name in self.issues_to_add:
            # prepare the new row
            issue: Issue = self.user.dict_issue[issue_name]
            add_row: int = self.new_row(issue.state)
            # add the new row
            self.excel_prop.insert_row(add_row)
            # add the values to the associated cells
            attrs = ["parent", "issue", "summary", "deadline"]
            values = [getattr(issue, attr) for attr in attrs]
            self._parse_attr_seq(attrs, values, add_row)
            TableCell(self.excel_prop, self.ws[f"C{add_row}"])

    def _parse_attr_seq(self, attrs: list[str], values: list[Any], row: int):
        """
        Add the attribute values to the table.

        :param attrs: the attribute names
        :type attrs: list[str]
        :param values: the attribute values
        :type values: list[Any]
        :param int row: the row in the table
        :return: None.
        """
        if len(attrs) != len(values):
            print("Improper lengths.")
            return
        for attr, value in zip(attrs, values):
            if attr not in self.dict_attr_column.keys():
                continue
            else:
                if value is not None:
                    column_letter: str = self.dict_attr_column[attr]
                    self.ws[f"{column_letter}{row}"].value = value

    def modify_all_issues(self):
        """Modify all issues in the table."""
        for issue_name in self.issues_to_modify:
            issue, table_cell = self.__get_issue_by_name(issue_name)
            row = table_cell.row
            # get the attributes to modify
            modify_attrs: list[str] = [attr for attr in self.attrs_column
                                       if getattr(table_cell, attr) != getattr(issue, attr)]
            # get the values to apply
            modify_values = [getattr(issue, attr) for attr in modify_attrs]
            # add the values to the table
            self._parse_attr_seq(modify_attrs, modify_values, row)
            # operate with the state changes separately
            if convert_issue_state(issue.state) != table_cell.state:
                self.modify_table_cell(issue.issue, issue.state)

    def modify_table_cell(self, issue_name: str, state: str):
        """
        Modify the table cell state.

        :param str issue_name: the issue name
        :param str state: the issue state
        :return: None.
        """
        table_cell: TableCell = self.table_cell_item(issue_name)
        row: int = table_cell.row
        add_row: int = self.new_row(state)
        self.excel_prop.insert_row(add_row)
        for cell in table_cell.cell_range():
            self.excel_prop.replace_cell(from_=cell, to_=self.ws[f"{cell.column_letter}{add_row}"])
        self.excel_prop.replace_cell(from_=f"NT{row}", to_=f"NT{add_row}")
        self.excel_prop.delete_row(row)

    def table_cell_item(self, issue_name: str) -> Optional[TableCell]:
        """
        Get the TableCell instance by the issue name.

        :param str issue_name: the issue name
        :return: the TableCell instance if exists.
        :rtype: TableCell or None
        """
        return self.excel_prop.table_cell_issue(issue_name)

    def __get_issue_by_name(self, issue_name: str) -> tuple[Issue, TableCell]:
        """
        Get the Issue and TableCell instances by the issue name.

        :param str issue_name: the issue name
        :return: the Issue and TableCell instances.
        :rtype: tuple[Issue, TableCell]
        """
        return self.user.dict_issue[issue_name], self.table_cell_item(issue_name)

    def add_new_work_items(self):
        """Add the new work items."""
        for issue_name in self.work_item_names:
            table_cell: TableCell = self.table_cell_item(issue_name)
            # get the work items from the YouTrack
            work_items_issue: list[IssueWorkItem] = self.user.dict_issue_work_item[issue_name]
            # get the work item dates from the YouTrack
            work_item_dates: set[datetime.date] = set(work_item.date for work_item in work_items_issue)
            # get the work item dates from the table
            table_cell_dates: set[datetime.date] = set(date for _, date, _ in table_cell.work_items())
            # get the work item dates to add
            new_work_items_dates: set[datetime.date] = work_item_dates.difference(table_cell_dates)
            # get the new work items
            new_work_items: list[IssueWorkItem] = [
                work_item for work_item in self.user.dict_issue_work_item.values()
                if work_item.date in new_work_items_dates]
            for work_item in new_work_items:
                work_item_tuple = work_item.issue, work_item.date, work_item.spent_time
                table_cell.add_work_item(work_item_tuple)

    def set_work_item_styles(self):
        """Set the styles to the work items."""
        for issue_name in self.user.issue_names():
            table_cell: TableCell = self.table_cell_item(issue_name)
            for work_item in table_cell.work_items():
                cell: Cell = table_cell.mapping_work_item_cell(work_item)
                style_name: str = table_cell.proper_work_item_style(work_item)
                self.excel_prop.styles.set_style(style_name, cell)

    def update_states(self):
        """Update the issue states in the table."""
        dict_updated_states: dict[str, str] = self.user.get_current_states(self.issues_to_update_state)
        for issue_name, state in dict_updated_states.items():
            if self.table_cell_item(issue_name).state != state:
                self.modify_table_cell(issue_name, state)


def old_main():
    youtrack_config = UserConfig().set_config_file(UserConfig.path)
    user = User(youtrack_config)
    user.pre_processing()
    # for issue_name, issue in user.dict_issue.items():
    #     print(f"issue_name = {issue_name}, issue = {repr(issue)}")
    # for issue_work_item_name, issue_work_item in user.dict_issue_work_item.items():
    #     print(f"issue_work_item_name = {issue_work_item_name}, issue_work_item = {repr(issue_work_item)}")

    path = Path(youtrack_config.get_json_attr("path_table")).resolve()
    name = path.name
    path_new_workbook = path.with_name(f"edited_{name}")

    wb: Workbook = openpyxl.load_workbook(path)
    ws_title = wb.sheetnames[1]
    ws: Worksheet = wb[ws_title]

    style_list = _StyleWorkItemList("styles")

    excel_prop = ExcelProp(ws, "excel_prop", style_list)
    # excel_prop.pre_processing()
    excel_prop.delete_row(12)
    for table_cell in excel_prop.dict_table_cell.values():
        table_cell.cell_hyperlink_nullify()
        table_cell.cell_hyperlink()
    # for table_cell in excel_prop.table_cells:
    #     print(table_cell.hyperlink)
    # join_user_xl = JoinUserXL(user, excel_prop)
    # join_user_xl.add_all_issues()
    # join_user_xl.modify_all_issues()
    # join_user_xl.add_new_work_items()
    # join_user_xl.set_work_item_styles()

    print("The work is finished.")
    wb.save(path_new_workbook)


def new_main():
    youtrack_config = UserConfig().set_config_file(UserConfig.path)
    user = User(youtrack_config)
    Issue(issue="DOC_ST-1467", state="Active", summary="Документация на ПДРА.465672.011 АРМ ДЛ МИ-2 ТИ",
          parent="DOC_ST-717", deadline=None, user=user)
    Issue(issue="DOC_ST-1611", state="Active", summary="Документы для поставки 021.093 с ВП", parent="None",
          deadline=datetime.date.fromisoformat("2022-04-26"), user=user)
    Issue(issue="DOC_ST-1589", state="Done", summary="Документация 021.096 (Сарапульский завод)", parent="None",
          deadline=datetime.date.fromisoformat("2022-04-25"), user=user)
    Issue(issue="DOC_ST-1649", state="Active", summary="Документы для 021.173", parent="None",
          deadline=datetime.date.fromisoformat("2022-05-06"), user=user)
    Issue(issue="DOC_ST-1735", state="Active", summary="Документы для 021.183 (АТС 48FXS, 16FXO, 1Е1 - 2 комплекта)",
          parent="None", deadline=datetime.date.fromisoformat("2022-06-06"), user=user)
    Issue(issue="DOC_ST-1478", state="Active", summary="Документы Севмаш 321.001, 321.002", parent="None",
          deadline=datetime.date.fromisoformat("2021-08-13"), user=user)
    Issue(issue="DOC_ST-1738", state="New", summary="Документы МКС-П 522 (38 комплектов) Отв. исполнитель Бровко",
          parent="None", deadline=datetime.date.fromisoformat("2022-07-10"), user=user)
    Issue(issue="DOC_ST-1588", state="New", summary="Документы для поставки 021.145", parent="None",
          deadline=datetime.date.fromisoformat("2022-03-25"), user=user)
    Issue(issue="DOC_ST-1313", state="Active",
          summary="МКС-П (ТИ) Комплект ВКС персональный ПДРА.465979.007-01 (на базе Гранат-П)", parent="DOC_ST-1252",
          deadline=datetime.date.fromisoformat("2020-12-14"), user=user)
    Issue(issue="DOC_ST-1553", state="Active",
          summary="Разработка ЭД для  мобильного комплекса связи Циркон ПДРА.465672.018", parent="None",
          deadline=datetime.date.fromisoformat("2021-11-10"), user=user)
    Issue(issue="DOC_ST-1720", state="Active", summary="Комплект ЭД на Циркон-В2 022.023", parent="None",
          deadline=datetime.date.fromisoformat("2022-05-20"), user=user)
    Issue(issue="DOC_ST-899", state="Active", summary="Комплект документации на изделие АТС-Протей ПДРА.465684.020",
          parent="DOC_ST-720", deadline=datetime.date.fromisoformat("2021-03-05"), user=user)
    Issue(issue="DOC-173", state="Active", summary="Отчеты, совещания и прочие административные задачи.",
          parent="DOC-22", deadline=None, user=user)
    Issue(issue="DOC_ST-1730", state="New", summary="Документы Севмаш 321.003 с ВП", parent="None",
          deadline=datetime.date.fromisoformat("2022-09-09"), user=user)
    Issue(issue="DOC_ST-896", state="Active",
          summary="Комплект эксплуатационной документации на изделие Сапфир-УТВ ПДРА.466533.010", parent="DOC_ST-720",
          deadline=datetime.date.fromisoformat("2021-03-05"), user=user)
    Issue(issue="DOC_ST-1224", state="New",
          summary="Инструкция по использованию ЗИП-О    ПДРА.465685.204И1 изм.2		  Корректировка комплекта ЭД на МКС-П для типовых испытаний",
          parent="DOC_ST-1215", deadline=datetime.date.fromisoformat("2020-10-20"), user=user)
    Issue(issue="DOC_ST-1220", state="Active",
          summary="Формуляр   ПДРА.465685.204ФО изм.3	Корректировка комплекта ЭД на МКС-П для типовых испытаний",
          parent="DOC_ST-1215", deadline=datetime.date.fromisoformat("2020-10-20"), user=user)
    Issue(issue="DOC_ST-1311", state="Active",
          summary="МКС-П (ТИ) Комплект ВКС персональный ПДРА.465979.007 (на базе Висмут-П)", parent="DOC_ST-1252",
          deadline=datetime.date.fromisoformat("2020-12-14"), user=user)
    Issue(issue="DOC_ST-1359", state="Active", summary="Комплект ЭД на ПДРА.465684.007-01 КУС-3 (ТИ)",
          parent="DOC_ST-717", deadline=datetime.date.fromisoformat("2021-01-31"), user=user)
    Issue(issue="DOC_ST-1653", state="Active", summary="Преобразователь 27-54, Гранат-С4К, МВЗ.", parent="None",
          deadline=datetime.date.fromisoformat("2022-02-11"), user=user)
    Issue(issue="DOC-1141", state="New", summary="Документация Заказчика 021.302", parent="None",
          deadline=datetime.date.fromisoformat("2022-02-28"), user=user)
    Issue(issue="DOC_ST-1592", state="Active",
          summary="Инструкция по антивирусной защите (код Д11) на изделия КОТС. Исполнитель: Бровко",
          parent="DOC_ST-720", deadline=datetime.date.fromisoformat("2021-11-25"), user=user)
    Issue(issue="DOC_ST-1306", state="Active",
          summary="Актуализировать документацию для сертификации комплекса оборудования \"Протей-imSwitch5 СП\" (версия ПО: 4.3)",
          parent="None", deadline=datetime.date.fromisoformat("2021-02-15"), user=user)
    Issue(issue="DOC_ST-1356", state="Active",
          summary="Сертификация «Протей-imSwitch5 СП» (ИК-2020): Документ ПДРА.4604021.050 005-2.0 РЭ Руководство по эксплуатации",
          parent="DOC_ST-1306", deadline=None, user=user)
    Issue(issue="DOC-180", state="New", summary="Обновление Технического описания продуктов SSW4/5", parent="None",
          deadline=datetime.date.fromisoformat("2021-12-05"), user=user)
    Issue(issue="ARCH_ST-89", state="New",
          summary="Оформление в архиве CD-диска с тестовыми конфигурационными файлами для проверки КПА Фобос",
          parent="DOC_ST-717", deadline=None, user=user)
    Issue(issue="DOC_ST-1423", state="Active", summary="Документация для сертификации ПК Гелиос R6",
          parent="DOC_ST-1370", deadline=datetime.date.fromisoformat("2021-03-26"), user=user)
    Issue(issue="ARCH_ST-87", state="New",
          summary="Подготовка учтенных дисков со сканами КД и ЭД на МКС-П для заказчика", parent="DOC_ST-717",
          deadline=datetime.date.fromisoformat("2021-08-30"), user=user)
    Issue(issue="DOC_ST-1349", state="Active",
          summary="Сертификация «Протей-imSwitch5 СП» (ИК-2020): Документ ПДРА.49010-02 13 Описание программы",
          parent="DOC_ST-1306", deadline=None, user=user)
    Issue(issue="DOC_ST-1354", state="Active",
          summary="Сертификация «Протей-imSwitch5 СП» (ИК-2020): Документ ПДРА.49010-02 97 Инструкция по созданию Диска восстановления",
          parent="DOC_ST-1306", deadline=None, user=user)
    Issue(issue="DOC_ST-1365", state="Active",
          summary="Поддержка в актуальном состоянии тестовой зоны отдела тех. документации", parent="None",
          deadline=None, user=user)
    Issue(issue="DOC-645", state="Active", summary="Актуализировать \"Руководство пользователя. CDR-файлы\" imSwitch5",
          parent="None", deadline=datetime.date.fromisoformat("2021-01-24"), user=user)
    Issue(issue="DOC-367", state="New", summary="Обновление документа mCore.MKD.ConfigFile.UserGuide", parent="None",
          deadline=datetime.date.fromisoformat("2020-11-29"), user=user)
    Issue(issue="DOC-82", state="Active", summary="Обновление документации на mGate.ITG", parent="None",
          deadline=datetime.date.fromisoformat("2020-11-29"), user=user)
    Issue(issue="ARCH-1", state="Active", summary="Учет архивной документации", parent="None",
          deadline=None, user=user)
    Issue(issue="ARCH-19", state="New", summary="Книга учета CD-дисков в архиве", parent="None",
          deadline=None, user=user)
    Issue(issue="DOC-672", state="Active", summary="Актуализация документации signalling", parent="None",
          deadline=datetime.date.fromisoformat("2020-07-17"), user=user)
    Issue(issue="DOC-353", state="Active", summary="Создание XSM.TAKT.645 User Guide", parent="None",
          deadline=datetime.date.fromisoformat("2020-03-31"), user=user)
    path = Path(youtrack_config.get_json_attr("path_table")).resolve()
    name = path.name
    path_new_workbook = path.with_name(f"edited_{name}")

    wb: Workbook = openpyxl.load_workbook(path)
    ws_title = wb.sheetnames[1]
    ws: Worksheet = wb[ws_title]

    style_list = _StyleWorkItemList("styles")

    excel_prop = ExcelProp(ws, "excel_prop", style_list)
    excel_prop._empty_rows()
    excel_prop.pre_processing()
    # excel_prop.delete_row(12)
    for table_cell in excel_prop.dict_table_cell.values():
        table_cell.cell_hyperlink_nullify()
        table_cell.cell_hyperlink()
    # print(excel_prop.table_cell_names())

    join_user_xl = JoinUserXL(user, excel_prop)
    join_user_xl.add_all_issues()
    print(excel_prop.table_cell_names())
    join_user_xl.modify_all_issues()
    join_user_xl.add_new_work_items()
    join_user_xl.set_work_item_styles()

    print("The work is finished.")
    wb.save(path_new_workbook)
    wb.close()


def main():
    path = r"/Users/user/Desktop/2022_template.xlsx"
    wb: Workbook = openpyxl.load_workbook(path)
    ws: Worksheet = wb["12 мес."]
    cell: Cell = ws["NS4"]
    print(cell.number_format)
    print(cell.data_type)
    print(cell.alignment)
    print(cell.border)
    print(cell.fill)
    print(cell.font)
    print(cell.protection)


if __name__ == "__main__":
    main()
