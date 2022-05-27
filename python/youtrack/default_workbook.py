import datetime
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter, coordinate_from_string, column_index_from_string, coordinate_to_tuple
from _style_work_item import _StyleWorkItemList


class ConstDefaultWs:
    """

    """
    # style_name: (rus_name, cell_style_coordinate, cell_legend_coordinate
    dict_legend = {
        'weekend': ('выходные', 'B21', 'C21'), 'deadline': ('дедлайн', 'B22', 'C22'), 'done': ('done', 'B23', 'C23'),
        'active': ('active', 'B24', 'C24'), 'test': ('test', 'B25', 'C25'),
        'going_start': ('планирую начать', 'B26', 'C26'), 'paused': ('paused', 'B27', 'C27'),
        'verified_closed': ('verified, closed', 'B28', 'C28'), 'going_finish': ('планирую завершить', 'B29', 'C29'),
        'sick': ('больничный', 'B30', 'C30'), 'vacation': ('отпуск', 'B31', 'C31')
    }
    # title_name: (rus_name, cell_title_coordinate)
    dict_titles = {
        'parent': ('РОД. ЗАДАЧА', 'B2'), 'name': ('ЗАДАЧА', 'C2'), 'summary': ('ОПИСАНИЕ', 'D2'),
        'deadline': ('DEADLINE', 'E2'), 'sum': ('Σ', 'NS2'), 'commentary': ('ПРИМЕЧАНИЕ', 'NT2')
    }
    # month_name: (rus_name, cell_title_coordinate, num_days)
    dict_month_ranges: dict[str, tuple[str, str, int]] = {
        'january': ('ЯНВАРЬ', 'F1', 31), 'february': ('ФЕВРАЛЬ', 'AL1', 28), 'march': ('МАРТ', 'BO1', 31),
        'april': ('АПРЕЛЬ', 'CU1', 30), 'may': ('МАЙ', 'DZ1', 31), 'june': ('ИЮНЬ', 'FF1', 30),
        'july': ('ИЮЛЬ', 'GK1', 31), 'august': ('АВГУСТ', 'HQ1', 31), 'september': ('СЕНТЯБРЬ', 'IW1', 30),
        'october': ('ОКТЯБРЬ', 'KB1', 31), 'november': ('НОЯБРЬ', 'LH1', 30), 'december': ('ДЕКАБРЬ', 'MM1', 31)
    }

    dict_month_titles: dict[str, tuple[str, str, str]] = {
        'january': ('ЯНВАРЬ', 'G2', 'AK2'), 'february': ('ФЕВРАЛЬ', 'AM2', 'BN2'), 'march': ('МАРТ', 'BP2', 'CT2'),
        'april': ('АПРЕЛЬ', 'CV2', 'DY2'), 'may': ('МАЙ', 'EA2', 'FE2'), 'june': ('ИЮНЬ', 'FG2', 'GJ2'),
        'july': ('ИЮЛЬ', 'GL2', 'HP2'), 'august': ('АВГУСТ', 'HR2', 'IV2'), 'september': ('СЕНТЯБРЬ', 'IX2', 'KA2'),
        'october': ('ОКТЯБРЬ', 'KC2', 'LG2'), 'november': ('НОЯБРЬ', 'LI2', 'ML2'),
        'december': ('ДЕКАБРЬ', 'MN2', 'NR2')
    }

    cells_headers: tuple[tuple[int, str]] = ((3, 'Active'), (7, 'New/Paused'), (11, 'Active/Done'), (15, 'Verified'))
    # holidays
    yy: int = datetime.date.today().year
    start_day = datetime.date(yy, 1, 1)
    end_day = datetime.date(yy + 1, 1, 1)
    time_delta = datetime.timedelta(days=1)
    holidays = {
        datetime.date(yy, 1, 1), datetime.date(yy, 1, 2), datetime.date(yy, 1, 3), datetime.date(yy, 1, 4),
        datetime.date(yy, 1, 5), datetime.date(yy, 1, 6), datetime.date(yy, 1, 7), datetime.date(yy, 2, 23),
        datetime.date(yy, 3, 7), datetime.date(yy, 3, 8), datetime.date(yy, 5, 1), datetime.date(yy, 5, 2),
        datetime.date(yy, 5, 3), datetime.date(yy, 5, 8), datetime.date(yy, 5, 9), datetime.date(yy, 5, 10),
        datetime.date(yy, 6, 12), datetime.date(yy, 6, 13), datetime.date(yy, 11, 4), datetime.date(yy, 12, 31)
    }
    day = start_day
    while day + time_delta <= end_day:
        if day.weekday() in (5, 6):
            holidays.add(day)
        day += time_delta

    weekend_rows = (4, 5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18)
    # for set_fill_color:
    #     # type_fill: str, fill_color: str, tint: float = None, theme: int = None
    dict_months_colors = {
        'january': ('tmp', 'FF0297CC', None, None), 'february': ('tmp', 'FF0BA6B6', None, None),
        'march': ('tmp', 'FF3AAA66', None, None), 'april': ('tmp', 'FF8BBD36', None, None),
        'may': ('tmp', 'FFD0CA04', None, None), 'june': ('tmp', 'FFF9AD01', None, None),
        'july': ('tmp', 'FFF08002', None, None), 'august': ('tmp', 'FFE94442', None, None),
        'september': ('tmp', 'FFCF687D', None, None), 'october': ('tmp', 'FF98668B', None, None),
        'november': ('tmp', 'FF697FB9', None, None), 'december': ('tmp', 'FF0078A9', None, None)
    }


def month_cell_periods() -> list[str]:
    """
    Merge cells for the month titles.

    :return: the merged cells.
    :rtype: list[str]
    """
    start_cells = [start for _, start, _ in ConstDefaultWs.dict_month_titles.values()]
    end_cells = [end for _, _, end in ConstDefaultWs.dict_month_titles.values()]
    return [f"{start_cell}:{end_cell}" for start_cell, end_cell in zip(start_cells, end_cells)]


def add_column(coord: str, add_value: int) -> str:
    """
    Find the cell in the _row that is N columns to the left/right.

    :param str coord: the base cell coordinate
    :param int add_value: the number of columns with the sign
    :return: the new cell coordinate.
    :rtype: str
    """
    column, row = coordinate_from_string(coord)
    col_idx = column_index_from_string(column)
    if add_value < 0 and col_idx + add_value < 1:
        print('The incorrect work of the add_column method.')
        return coord
    add_col_idx = col_idx + add_value
    add_column_letter = get_column_letter(add_col_idx).upper()
    return f'{add_column_letter}{row}'


def range_coord(start_coord: str, end_coord: str) -> tuple[int, int, int, int]:
    """
    Convert the range string with the start and end coordinates to the tuple:\n
    (min_row, min_col, max_row, max_col)\n

    :param str start_coord: the start cell coordinates
    :param str end_coord: the end cell coordinates
    :return: the values of min_row, min_col, max_row, max_col.
    :rtype: tuple[int, int, int, int]
    """
    # convert start coordinate
    start_row, start_column = coordinate_to_tuple(start_coord)
    # convert end coordinate
    end_row, end_column = coordinate_to_tuple(end_coord)
    # find min and max values
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)
    min_col = min(start_column, end_column)
    max_col = max(start_column, end_column)
    return min_row, max_row, min_col, max_col


class WorksheetDefault:
    """
    Specify the default worksheet.

    Properties:
        weekend_columns --- specify the column letters for the weekend style;\n

    Functions:
        set_dates() --- specify the dates in row 1;\n
        set_month_names() --- specify the month names;\n
        month_cell_merge() --- merge cells for the month title;\n
        set_month_titles() --- specify the month titles;\n
        set_default() --- apply the basic style for the whole table;\n
        set_weekends() --- apply the weekend style;\n
        set_headers() --- apply the headers style to the headers;\n
        set_headers_row() --- apply the style to the headers rows;\n
        set_title() --- apply the title style to the title row;\n
        set_legend() --- specify the legend;\n
    """
    def __init__(self, wb: Workbook, ws: Worksheet, styles: _StyleWorkItemList):
        self.wb = wb
        self.ws = ws
        self.styles = styles

    def __str__(self):
        return f"Workbook: {self.wb}, Worksheet: {self.ws}"

    def __repr__(self):
        return f"WorksheetDefault(wb={self.wb}, ws={self.ws})"

    def __hash__(self):
        return hash((self.wb.properties.name, self.ws.title))

    def __key(self):
        return self.wb.properties.name, self.ws.title

    def __eq__(self, other):
        if isinstance(other, WorksheetDefault):
            return self.__key() == other.__key()
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, WorksheetDefault):
            return self.__key() != other.__key()
        else:
            return NotImplemented

    def __contains__(self, item):
        return item in self.styles

    def __cell_range(self, start_coord: str, end_coord: str):
        """
        Converts the cell range to the cell generator in the range.

        :param str start_coord: the start cell coordinates
        :param str end_coord: the end cell coordinates
        :return: the generator of cells in the cell range.
        """
        min_row, max_row, min_col, max_col = range_coord(start_coord, end_coord)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                yield self.ws[f'{get_column_letter(col)}{row}']

    def set_dates(self):
        """Specify the dates in _row 1."""
        for index, month_params in enumerate(ConstDefaultWs.dict_month_ranges.values()):
            _, coord, num_days = month_params
            start_col_idx = coordinate_to_tuple(coord)[1]
            year = datetime.date.today().year
            month = index + 1

            for day in range(1, num_days + 1):
                col_idx = start_col_idx + day
                cell: Cell = self.ws[f'{get_column_letter(col_idx)}1']
                cell.value = datetime.date(year=year, month=month, day=day)
                self.styles.set_style("month_date", cell)

    @property
    def weekend_columns(self) -> list[str]:
        """
        Specify the column letters for the weekend style.

        :return: the column letters.
        :rtype: list[str]
        """
        return [cell.column_letter for cell in self.__cell_range("G1", "NR1") if cell.value in ConstDefaultWs.holidays]

    def set_month_names(self):
        """Specify the month names."""
        for month, month_params in ConstDefaultWs.dict_month_ranges.items():
            rus_name, coord, _ = month_params
            cell: Cell = self.ws[f"{coord}"]
            self.ws[f'{coord}'].value = rus_name
            self.styles.set_style(f"{month}_header", cell)

    def month_cell_merge(self):
        """Merge cells for the month title."""
        for period in month_cell_periods():
            self.ws.merge_cells(period)

    def set_month_titles(self):
        """Specify the month titles."""
        for month, month_values in ConstDefaultWs.dict_month_titles.items():
            rus_name, coord, _ = month_values
            cell: Cell = self.ws[f'{coord}']
            cell.value = rus_name
            self.styles.set_style(f"{month}_title", cell)

    def set_default(self):
        """Apply the basic style for the whole table."""
        for cell in self.__cell_range("A1", "NT18"):
            self.styles.set_style("basic", cell)

    def set_weekends(self):
        """Apply the weekend style."""
        for column_letter in self.weekend_columns:
            for row in ConstDefaultWs.weekend_rows:
                cell: Cell = self.ws[f'{column_letter}{row}']
                self.styles.set_style("weekend", cell)

    def set_headers(self):
        """Apply the headers style to the headers."""
        for row, value in ConstDefaultWs.cells_headers:
            # merge the cells
            self.ws.merge_cells(f'B{row}:E{row}')
            cell: Cell = self.ws[f'B{row}']
            cell.value = value
            self.styles.set_style("header", cell)

    def set_headers_row(self):
        """Apply the style to the headers rows."""
        for row, _ in ConstDefaultWs.cells_headers:
            start_coord = f'F{row}'
            end_coord = f'NT{row}'
            for cell in self.__cell_range(start_coord, end_coord):
                self.styles.set_style("header", cell)

    def set_title(self):
        """Apply the title style to the title _row."""
        for title, coord in ConstDefaultWs.dict_titles.values():
            cell: Cell = self.ws[f'{coord}']
            cell.value = title
            self.styles.set_style("title", cell)

    def set_legend(self):
        """Specify the legend."""
        self.ws.merge_cells('B20:C20')
        self.ws['B20'].value = 'Легенда'
        for style_name, values in ConstDefaultWs.dict_legend.items():
            rus_name, coord, legend = values
            cell: Cell = self.ws[f"{coord}"]
            self.styles.set_style(style_name, cell)
            self.ws[f"{legend}"].value = rus_name


def main():
    pass


if __name__ == '__main__':
    main()
