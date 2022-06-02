from collections import Counter
from typing import Optional, Union, Any
import datetime
import numpy
from numpy import ndarray, busdaycalendar
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, column_index_from_string
from decimal import Decimal
from _style_work_item import _StyleWorkItemList
from copy import copy


def check_coord(coord: str) -> bool:
    """
    Verify the cell coordinate.

    :param str coord: the cell coordinates
    :return: the verification flag.
    :rtype: bool
    """
    flag = False
    try:
        coordinate_to_tuple(coord)
    except TypeError as e:
        print(f'TypeError {str(e)}. Incorrect value type.\n')
        print(f'coordinate = {coord}')
    except ValueError as e:
        print(f'ValueError {str(e)}. Incorrect value.\n')
        print(f'coordinate = {coord}')
    except OSError as e:
        print(f'OSError {e.errno}, {e.strerror}.\n')
        print(f'coordinate = {coord}')
    else:
        flag = True
    finally:
        return flag


def range_coord(start_coord: str, end_coord: str):
    """
    Convert the range string with the start and end coordinates to the tuple:\n
    (min_row, min_col, max_row, max_col)\n

    :param start_coord: the start cell coordinate, str
    :param end_coord: the end cell coordinate, str
    :return: the values for iteration of the tuple[int, int, int, int] type.
    """
    # convert start coordinate
    start_row, start_column = coordinate_to_tuple(start_coord)
    # convert end coordinate
    end_row, end_column = coordinate_to_tuple(end_coord)
    # find min and max values
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)

    min_col = min(start_column, end_column)
    max_col = max(start_column, end_column)
    return min_row, max_row, min_col, max_col


def convert_spent_time(spent_time: Any) -> Decimal:
    """
    Convert the spent time to the specified format.

    :param spent_time: the converted value.
    :return: the decimal value.
    :rtype: Decimal
    """
    return Decimal(spent_time).normalize()


def convert_datetime_date(value: Any) -> Optional[datetime.date]:
    """
    Convert the datetime or None to the date.

    :param value: the value to convert
    :return: the date value.
    :rtype: datetime.date or None
    """
    if isinstance(value, datetime.date):
        return value
    elif isinstance(value, datetime.datetime):
        return value.date()
    else:
        return None


def calendar_days() -> tuple[list[datetime.date], list[datetime.date]]:
    """
    Get the business day and weekend dates.

    :return: the date separated by the type.
    :rtype: tuple[list[datetime.date], list[datetime.date]]
    """
    # the current year
    year: int = datetime.date.today().year
    # the weekend days
    holidays: set[datetime.date] = {
        datetime.date(year, 1, 1), datetime.date(year, 1, 2), datetime.date(year, 1, 3), datetime.date(year, 1, 4),
        datetime.date(year, 1, 5), datetime.date(year, 1, 6), datetime.date(year, 1, 7), datetime.date(year, 2, 23),
        datetime.date(year, 3, 7), datetime.date(year, 3, 8), datetime.date(year, 5, 1), datetime.date(year, 5, 2),
        datetime.date(year, 5, 3), datetime.date(year, 5, 8), datetime.date(year, 5, 9), datetime.date(year, 5, 10),
        datetime.date(year, 6, 12), datetime.date(year, 6, 13), datetime.date(year, 11, 4), datetime.date(year, 12, 31)
    }
    # convert to the ISO
    holidays_iso: list[str] = [date_holiday.isoformat() for date_holiday in holidays]
    # the date range
    start_day = datetime.date(year, 1, 1)
    end_day = datetime.date(year + 1, 1, 1)
    # specify the date range
    year_days: ndarray = numpy.arange(numpy.datetime64(start_day.isoformat()), numpy.datetime64(end_day.isoformat()))
    # convert to list
    year_days_list: list[datetime.date] = year_days.tolist()
    # specify the business day calendar
    business_cal: busdaycalendar = numpy.busdaycalendar(holidays=holidays_iso)
    # get the bool values if the date is a business one
    is_bus_days: ndarray = numpy.is_busday(dates=year_days, busdaycal=business_cal)
    # convert to list
    is_bus_days_list: list[bool] = is_bus_days.tolist()

    bus_days = [bus_day for bus_day, is_business in zip(year_days_list, is_bus_days_list) if is_business]
    week_ends = [week_day for week_day in set(year_days_list).difference(set(bus_days))]
    return bus_days, week_ends


class ExcelProp:
    """
    Define the Excel Worksheet properties.

    Constants:
        dict_headers --- the dictionary of the states and indexes;\n
        dict_headers_short --- the dictionary of the issue states and indexes;\n
        dict_state_style --- the dictionary of the states and cell styles;\n
        dict_state_priority --- the dictionary of the states and priorities;\n
        month_columns --- the columns with the month titles;\n

    Class params:
        dict_table_cell --- the dictionary of the TableCell instances;\n

    Params:
        ws --- the worksheet;\n
        name --- the instance name;\n
        styles --- the styles implemented into the worksheet;\n

    Properties:
        headers --- the header cells;\n
        headers_short --- the header cells without the legend;\n
        headers_row --- the header cell rows;\n
        headers_row_short --- the header cell rows without the legend;\n
        cell_states --- the dictionary of the cell coordinates and states;\n
        table_cells --- the cells for the TableCell instances;\n
        table_cell_names --- the issue names;\n
        bus_columns -- the work day date columns;\n
        weekend_columns --- the weekend date columns;\n


    Functions:
        get_column_by_date(date) --- get the column letter for the specified date;\n
        cell_in_range(start_coord, end_coord) --- generate the cells in the range;\n
        delete_row(row) --- delete the row from the table;\n
        insert_row(row) --- add the row to the table;\n
        _separate_headers() --- unmerge the header cells;\n
        _combine_headers() --- merge the header cells;\n
        _cells_formulae() --- specify the formulae to the cells for the total spent time;\n
        _check_empty(item) --- verify if the Cell, coordinate, or row is empty;\n
        _empty_rows() --- delete all empty rows except for the pre-headers;\n
        replace_cell(*, from_, to_) --- replace the cell attribute values to another cell;\n
        add_work_item(work_item) --- add the work item to the table;\n
        get_date_by_cell(cell) --- get the date associated with the cell;\n
        __index(state) --- get the state index;\n
        pre_processing() --- pre-process the table to get rid of empty rows and repeating issues;\n
        table_cell_issue(table_base) --- get the TableCell instance based on the row, cell, or issue name;\n
        default_row(row) --- specify the default row;\n
        update_header() --- specify the style to the header rows
    """
    dict_headers = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3, 'Легенда': 4}
    dict_headers_short = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3}
    dict_state_style = {
        "New": "basic",
        "New/Paused": "paused",
        "Active": "active",
        "Paused": "paused",
        "Active/Paused": "active",
        "Done": "done",
        "Test": "test",
        "Done/Test": "done",
        "Verified": "verified_closed",
        "Closed": "verified_closed",
        "Discuss": "paused",
        "Review": "paused",
        "Canceled": "verified_closed"
    }
    month_columns = ("F", "AL", "BO", "CU", "DZ", "FF", "GK", "HQ", "IW", "KB", "LH", "MM")
    dict_state_priority: dict[str, int] = {"Active": 30, "New/Paused": 10, "Done/Test": 20, "Verified": 40}
    dict_table_cell = dict()

    def __init__(self, ws: Worksheet, name: str, styles: _StyleWorkItemList):
        self.ws = ws
        self.name = name
        self.styles = styles

    def __str__(self):
        return f"Worksheet: {self.ws.title}, name: {self.name}"

    def __repr__(self):
        return f"ExcelProp(ws={self.ws}, name={self.name})"

    def __format__(self, format_spec):
        if format_spec == "base":
            return str(self)
        elif format_spec == "row_equal":
            return repr(self)
        else:
            return f"{self.name}"

    def __hash__(self):
        return hash((self.ws, self.name))

    def __key(self):
        return self.ws, self.name

    def __eq__(self, other):
        if isinstance(other, ExcelProp):
            return self.__key() == other.__key()
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, ExcelProp):
            return self.__key() != other.__key()
        else:
            return NotImplemented

    def get_column_by_date(self, date: datetime.date) -> str:
        """
        Get the column letter for the specified date.

        :param date: the date
        :type date: datetime.date
        :return: the column_letter.
        :rtype: str
        """
        cell: Cell
        for cell in self.cell_in_range("G1", "NR1"):
            if convert_datetime_date(cell.value) == date:
                return cell.column_letter

    def cell_in_range(self, start_coord: str, end_coord: str):
        """
        Convert the cell range to the cell generator in the range.

        :param str start_coord: the start cell coordinates
        :param str end_coord: the end cell coordinates
        :return: the generator of cells.
        """
        if check_coord(coord=start_coord) and check_coord(coord=end_coord):
            min_row, max_row, min_col, max_col = range_coord(start_coord=start_coord, end_coord=end_coord)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    coord = f'{get_column_letter(col)}{row}'
                    cell: Cell = self.ws[coord]
                    yield cell

    @property
    def headers(self) -> list[Cell]:
        """
        Get the header cells.

        :return: the cells.
        :rtype: list[Cell]
        """
        end_coord = f"B{self.ws.max_row}"
        return [cell for cell in self.cell_in_range("B3", end_coord) if cell.value in ExcelProp.dict_headers.keys()]

    @property
    def headers_short(self) -> list[Cell]:
        """
        Get the header cells without the legend.

        :return: the cells.
        :rtype: list[Cell]
        """
        return self.headers[:4]

    @property
    def headers_row(self) -> list[int]:
        """
        Get the header row values.

        :return: the list of header rows.
        :rtype: list[int]
        """
        return [header.row for header in self.headers]

    @property
    def headers_row_short(self) -> list[int]:
        """
        Get the header row values without the legend.

        :return: the list of header rows.
        :rtype: list[int]
        """
        return self.headers_row[:4]

    def delete_row(self, row: int):
        """
        Delete the row from the table.

        :param int row: the table row number
        :return: None.
        """
        self._separate_headers()
        table_cell: TableCell
        for table_cell in self.dict_table_cell.values():
            table_cell.cell_hyperlink_nullify()
        self.ws.delete_rows(row)
        self._combine_headers()
        self._cells_formulae()
        for table_cell in self.dict_table_cell.values():
            table_cell.cell_hyperlink()

    def insert_row(self, row: int):
        """
        Add the row to the table.

        :param int row: the table row number
        :return: None.
        """
        self._separate_headers()
        for table_cell in self.dict_table_cell.values():
            table_cell.cell_hyperlink_nullify()
        self.ws.insert_rows(row)
        self.default_row(row)
        self._combine_headers()
        self._cells_formulae()
        for table_cell in self.dict_table_cell.values():
            table_cell.cell_hyperlink()

    def _separate_headers(self):
        """Unmerge the header cells."""
        # the common headers
        for header_row in self.headers_row_short:
            start_coord = f"B{header_row}"
            end_coord = f"E{header_row}"
            self.ws.unmerge_cells(f"{start_coord}:{end_coord}")
            for cell_header in self.cell_in_range(start_coord, end_coord):
                cell_header.style = "basic"
        # the legend header
        header_legend_row = self.headers_row[4]
        start_legend_coord = f"B{header_legend_row}"
        end_legend_coord = f"C{header_legend_row}"
        self.ws.unmerge_cells(f"{start_legend_coord}:{end_legend_coord}")
        for cell_legend in self.cell_in_range(start_legend_coord, end_legend_coord):
            cell_legend.style = "basic"

    def _combine_headers(self):
        """Merge the header cells."""
        # the common headers
        for header_row in self.headers_row_short:
            self.ws.merge_cells(f"B{header_row}:E{header_row}")
            self.update_header()
        # the legend header
        header_legend_row = self.headers_row[4]
        self.ws.merge_cells(f"B{header_legend_row}:C{header_legend_row}")

    def _cells_formulae(self):
        """Set the formulae to the cells for the total spent time."""
        for cell in self.cell_in_range(f"NS{self.headers_row[0] + 1}", f"NS{self.headers_row[4]}"):
            row = cell.row
            if row not in self.headers_row:
                cell.value = f"=SUM(G{row}:NR{row})"
                self.styles.set_style("sum", cell)
            else:
                continue

    def _check_empty(self, item: Union[int, str, Cell]) -> bool:
        """
        Verify if the cell, cell coordinates, or cell row is empty.

        :param item: the cell value
        :type item: int or str or Cell
        :return: the emptiness flag.
        :rtype: bool
        """
        if isinstance(item, Cell):
            return True if item.value is None else False
        elif isinstance(item, str):
            return True if self.ws[f"{item}"].value is None else False
        elif isinstance(item, int):
            cell_values = (
                self.ws[f"B{item}"].value, self.ws[f"C{item}"].value,
                self.ws[f"D{item}"].value, self.ws[f"E{item}"].value)
            if any(value is not None for value in cell_values):
                return False
            else:
                return True

    def _empty_rows(self):
        """Delete all empty rows except for the pre-headers."""
        list_empty = []
        for state in self.dict_headers_short.keys():
            low_limit = self.headers_row[self.__index(state)] + 1
            if self.__index(state) != 3:
                high_limit = self.headers_row[self.__index(state) + 1] - 1
            else:
                high_limit = self.headers_row[self.__index(state) + 1] - 3
            for row in range(low_limit, high_limit):
                if self._check_empty(row):
                    list_empty.append(row)
        for empty_row in sorted(list_empty, reverse=True):
            self.delete_row(empty_row)

    def replace_cell(self, *, from_: Union[Cell, str], to_: Optional[Union[Cell, str]] = None):
        """
        Replace the cell attribute values to another cell. If to_ is None, the values are deleted.

        :param from_: the original cell or cell coordinates
        :type from_: Cell or str
        :param to_: the destination cell or cell coordinates
        :type to_: Cell or str or None
        :return: None.
        """
        cell_proxy: Optional[Cell]
        cell_base: Cell
        # if the cell coordinates
        if isinstance(from_, str):
            cell_base = self.ws[f"{from_}"]
        # if the cell
        else:
            cell_base = from_
        # if the destination cell exists
        if to_ is not None:
            if isinstance(to_, str):
                cell_proxy = self.ws[f"{to_}"]
            else:
                cell_proxy = to_
            # copy all required attribute values
            # if the destination cell is empty
            if cell_proxy.value is None:
                cell_proxy.value = copy(cell_base.value)
            # if the destination cell has some value
            else:
                cell_proxy.value += cell_base.value
            if cell_base.has_style:
                self.styles.set_style(cell_base.style, cell_proxy)
            else:
                cell_proxy.number_format = copy(cell_base.number_format)
                cell_proxy.style = "basic"
            # cell_proxy._style = copy(cell_base.style)
        # set the default cell attribute values
        cell_base.value = None
        cell_base.style = "basic"

    def add_work_item(self, work_item: tuple[str, datetime.date, Union[int, Decimal]]):
        """
        Add the work item to the table.

        issue_name: str, date: datetime.date, spent_time: Union[int, Decimal]
        :param work_item: the work item to add
        :type work_item: tuple[str, datetime.date, Decimal or int]
        :return: None.
        """
        issue_name, *_ = work_item
        self.dict_table_cell[issue_name].add_work_item(work_item)

    def get_date_by_cell(self, cell: Cell) -> Optional[datetime.date]:
        """
        Get the date associated with the cell.

        :param cell: the table cell
        :type cell: Cell
        :return: the date.
        :rtype: datetime.date or None
        """
        column = cell.column_letter
        raw_date = self.ws[f"{column}1"].value
        return convert_datetime_date(raw_date)

    def __index(self, state: str) -> int:
        """
        Get the state index.

        :param str state: the state
        :return: the header index.
        :rtype: int
        """
        return self.dict_headers_short[state]

    @property
    def cell_states(self):
        """
        Get the dictionary of the cell coordinates and states.

        :return: the cell coordinates and state mapping.
        :rtype: dict[str, str]
        """
        dict_cell_states: dict[str, str] = dict()
        for state in self.dict_headers_short.keys():
            start_row = self.headers_row[self.__index(state)] + 1
            end_row = self.headers_row[self.__index(state) + 1]
            start_coord = f"C{start_row}"
            end_coord = f"C{end_row}"
            for cell in self.cell_in_range(start_coord, end_coord):
                dict_cell_states[cell.coordinate] = state
        return dict_cell_states

    @property
    def table_cells(self) -> list[Cell]:
        """
        Get the cells for the TableCell instances.

        :return: the cells.
        :rtype: list[Cell]
        """
        return [self.ws[f"{coord}"] for coord in self.cell_states.keys() if self.ws[f"{coord}"].value is not None]

    def table_cell_issue(self, table_base: Union[int, Cell, str]):
        """
        Get the TableCell instance based on the row, cell, or issue name.

        :param table_base: the object to link to the TableCell instance
        :type table_base: int or Cell or str
        :return: the TableCell instance.
        :rtype: TableCell or None
        """
        if isinstance(table_base, int):
            cell = self.ws[f"C{table_base}"]
            return TableCell(self, cell)
        elif isinstance(table_base, str):
            return self.dict_table_cell[table_base]
        elif isinstance(table_base, Cell):
            return TableCell(self, table_base)
        else:
            print(f"Something went wrong. Table base: {table_base}, type: {type(table_base)}.")
            return None

    def table_cell_names(self) -> set[str]:
        """
        Get the issue names.

        :return: the issue names.
        :rtype: set[str]
        """
        return {cell.value for cell in self.table_cells}

    def pre_processing(self):
        """Pre-process the table to get rid of empty rows and repeating issues."""
        counter = Counter(self.table_cell_names())
        # get non-unique issues
        non_unique = [key for key, value in counter.items() if value > 1]
        for issue in non_unique:
            row_eq = _RowEqual(self, issue)
            row_eq.join_cells()
        # delete empty rows
        self._empty_rows()
        # initiate the TableCell instances
        for cell in self.table_cells:
            TableCell(self, cell)

    @property
    def bus_columns(self) -> list[str]:
        """
        Specify the business day columns.

        :return: the business day columns.
        :rtype: list[str]
        """
        return [self.get_column_by_date(date) for date in calendar_days()[0]]

    @property
    def weekend_columns(self) -> list[str]:
        """
        Specify the weekend columns.

        :return: the weekend columns.
        :rtype: list[str]
        """
        return [self.get_column_by_date(date) for date in calendar_days()[1]]

    def default_row(self, row: int):
        """Specify the default row when added."""
        # the basic
        for column in ("B", "C", "D", "NT"):
            self.ws[f"{column}{row}"].style = "basic_issue"
        # the cells with dates
        for column in self.bus_columns:
            self.ws[f"{column}{row}"].style = "basic"
        # the weekends
        for column in self.weekend_columns:
            self.ws[f"{column}{row}"].style = "weekend"
        # the business days
        for column in self.month_columns:
            self.styles.set_style("header_no_border", self.ws[f"{column}{row}"])
        # the issue deadline
        self.styles.set_style("deadline_issue", self.ws[f"E{row}"])
        # the total spent time cell
        self.styles.set_style("sum", self.ws[f"NT{row}"])

    def update_header(self):
        """Specify the style to the header rows."""
        for row in self.headers_row_short:
            self.styles.set_style("header_text", self.ws[f"B{row}"])
            for cell in self.cell_in_range(f"F{row}", f"NR{row}"):
                self.styles.set_style("header", cell)
            for column in self.month_columns:
                self.styles.set_style("header_no_border", self.ws[f"{column}{row}"])


class TableCell:
    """
    Define the table row.

    Params:
        excel_prop --- the ExcelProp instance;\n
        cell --- the base cell;\n
        issue --- the base cell value;\n

    Properties:
        ws --- the worksheet;\n
        row --- the row;\n
        state --- the state;\n
        cell_last --- the last work item cell;\n
        work_item_last --- the last work item;\n

    Functions:
        _shift_col_idx(column) --- get the shift between the column and the base cell;\n
        _shift_cell(shift) --- get the shifted cell in the same row;\n
        _cell_range() --- generate the cells in the range;\n
        _pyxl_cells() --- get the work item cells;\n
        work_items() --- get the work items;\n
        mapping_cell_work_item(cell) --- convert the cell to the work item;\n
        mapping_work_item_cell(work_item) --- convert the work item to the cell;\n
        add_work_item(work_item) --- add the work item to the row;\n
        proper_cell_style(cell) --- get the cell style based on the state;\n
        proper_work_item_style(work_item) --- get the work item style based on the state;\n
        cell_hyperlink() --- specify the hyperlinks;\n
        cell_hyperlink_nullify() --- delete the hyperlinks;\n
    """
    def __init__(self, excel_prop: ExcelProp, cell: Cell):
        self.excel_prop = excel_prop
        self.cell = cell
        self.issue = str(cell.value)

        self.excel_prop.dict_table_cell[self.issue] = self

    def __str__(self):
        return f"Table cell: {self.cell.coordinate}"

    def __repr__(self):
        return f"TableCell({self.excel_prop}, {self.cell})"

    def __hash__(self):
        return hash((self.excel_prop.name, self.cell.coordinate, self.issue))

    def __key(self):
        return self.excel_prop.name, self.cell.coordinate, self.issue

    def __eq__(self, other):
        if isinstance(other, TableCell):
            return self.__key() == other.__key()
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, TableCell):
            return self.__key() != other.__key()
        else:
            return NotImplemented

    def __lt__(self, other):
        if isinstance(other, TableCell):
            return self.row < other.row
        else:
            return NotImplemented

    def __gt__(self, other):
        if isinstance(other, TableCell):
            return self.row > other.row
        else:
            return NotImplemented

    def __le__(self, other):
        if isinstance(other, TableCell):
            return self.row <= other.row
        else:
            return NotImplemented

    def __ge__(self, other):
        if isinstance(other, TableCell):
            return self.row >= other.row
        else:
            return NotImplemented

    @property
    def ws(self) -> Worksheet:
        """
        Get the ExcelProp worksheet.

        :return: the worksheet.
        :rtype: Worksheet
        """
        return self.excel_prop.ws

    @property
    def row(self) -> int:
        """
        Get the base cell row.

        :return: the cell row.
        :rtype: int
        """
        return self.cell.row

    def _shift_col_idx(self, column: str) -> int:
        """
        Get the difference between the column and the base cell.

        :param str column: the column letter
        :return: the column shift.
        :rtype: int
        """
        return column_index_from_string(column) - self.cell.col_idx

    def _shift_cell(self, shift: int) -> Optional[Cell]:
        """
        Get the shifted cell in the same row.

        :param int shift: the cell shift
        :return: the shifted cell.
        :rtype: Cell or None
        """
        col_idx = self.cell.col_idx + shift
        if col_idx >= 1:
            column_letter = get_column_letter(col_idx)
            return self.ws[f"{column_letter}{self.row}"]
        else:
            print(f"ValueError, the shift {shift} is out of bounds.")
            return None

    @property
    def state(self) -> str:
        """
        Get the issue state.

        :return: the state.
        :rtype: str
        """
        return self.excel_prop.cell_states[self.cell.coordinate]

    @property
    def deadline(self) -> Optional[datetime.date]:
        """
        Get the issue deadline if exists.

        :return: the deadline.
        :rtype: datetime.date or None
        """
        return self._shift_cell(2).value

    @property
    def parent(self):
        """
        Get the parent issue name if exists.

        :return: the parent issue name.
        :rtype: str
        """
        return self._shift_cell(-1).value

    def cell_range(self):
        """
        Specify the cell generator for the range.

        :return: iterating through the cells.
        """
        return self.excel_prop.cell_in_range(f"G{self.row}", f"NR{self.row}")

    def _pyxl_cells(self) -> list[Cell]:
        """
        Get the work item cells.

        :return: the cells.
        :rtype: list[Cell]
        """
        return [cell for cell in self.cell_range() if cell.value is not None]

    def work_items(self) -> list[tuple[str, datetime.date, Decimal]]:
        """
        Get the work items.

        :return: the work items.
        :rtype: list[tuple[str, datetime.date, Decimal]]
        """
        return [self.mapping_cell_work_item(cell) for cell in self._pyxl_cells()]

    @property
    def cell_last(self) -> Cell:
        """
        Get the last work item cell.

        :return: the last work item cell.
        :rtype: Cell
        """
        max_column = max(cell.column_letter for cell in self._pyxl_cells())
        return self.ws[f"{max_column}{self.row}"]

    @property
    def work_item_last(self) -> tuple[str, datetime.date, Decimal]:
        """
        Get the last work item.

        :return: the work item.
        :rtype: tuple[str, datetime.date, Decimal]
        """
        cell: Cell = self.cell_last
        return self.mapping_cell_work_item(cell)

    def mapping_cell_work_item(self, cell: Cell) -> tuple[str, datetime.date, Decimal]:
        """
        Convert the cell to the work item.

        :param cell: the cell
        :type cell: Cell
        :return: the work item.
        :rtype: tuple[str, datetime.date, Decimal]
        """
        date = self.excel_prop.get_date_by_cell(cell)
        return self.issue, date, Decimal(cell.value).normalize()

    def mapping_work_item_cell(self, work_item: tuple[str, datetime.date, Decimal]) -> Cell:
        """
        Convert the work item to the cell.

        :param work_item: the work item
        :type work_item: tuple[str, datetime.date, Decimal]
        :return: the cell.
        :rtype: Cell
        """
        _, date, _ = work_item
        column = self.excel_prop.get_column_by_date(date)
        return self.ws[f"{column}{self.row}"]

    def add_work_item(self, work_item: tuple[str, datetime.date, Decimal]):
        """
        Add the work item to the row.

        :param work_item: the work item
        :type work_item: tuple[str, datetime.date, Decimal]
        :return: None.
        """
        _, date, spent_time = work_item
        cell = self.mapping_work_item_cell((self.issue, date, spent_time))
        cell.data_type = "n"
        cell.value = spent_time

    def proper_cell_style(self, cell: Cell) -> str:
        """
        Get the cell style based on the state.

        :param cell: the cell
        :type cell: Cell
        :return: the style name.
        :rtype: str
        """
        if cell not in self._pyxl_cells():
            return cell.style.name
        elif self.excel_prop.get_date_by_cell(cell) == self.deadline:
            return "deadline"
        elif cell != self.cell_last:
            return "active"
        elif self.state is not None:
            return self.excel_prop.dict_state_style[self.state]
        else:
            return "basic"

    def proper_work_item_style(self, work_item: tuple[str, datetime.date, Decimal]) -> str:
        """
        Get the proper work item style based on the state.

        :param work_item: the work item
        :type work_item: tuple[str, datetime.date, Decimal]
        :return: the style name
        :rtype: str
        """
        _, date, spent_time = work_item
        cell = self.mapping_work_item_cell((self.issue, date, spent_time))
        return self.proper_cell_style(cell)

    def cell_hyperlink(self):
        """Specify the issue and parent issue hyperlinks."""
        self.cell.hyperlink = f"https://youtrack.protei.ru/issue/{self.issue}" if self.issue is not None else None
        self._shift_cell(-1).hyperlink = f"https://youtrack.protei.ru/issue/{self.parent}" \
            if self.parent is not None else None

    def cell_hyperlink_nullify(self):
        """Discard the issue and parent issue hyperlinks."""
        self.cell.hyperlink = None
        self._shift_cell(-1).hyperlink = None


class _RowEqual:
    """
    Specify the instance to work with the issues having the same names.

    Params:
        excel_prop --- the ExcelProp instance;\n
        issue --- the issue name;\n

    Properties:
        cells --- the cells with the same issue names;\n
        dict_issues --- the dictionary of the issue cells and state priorities;\n

    Functions:
        work_item_cells() --- get the cells to join;\n
        cell_final() --- get the final cell to keep in the table;\n
        join_cells() --- join the cells into the one issue;\n
    """
    def __init__(self, excel_prop: ExcelProp, issue: str):
        self.excel_prop = excel_prop
        self.issue = issue

    def __str__(self):
        cell_string = ", ".join([cell.coordinate for cell in self.cells])
        return f"_RowEqual: issue name - {self.issue},\ncells - {cell_string}"

    def __repr__(self):
        return f"_RowEqual({format(self.excel_prop, 'row_equal')}, {self.issue})"

    def __len__(self):
        return len(self.cells)

    def __bool__(self):
        return len(self.cells) > 1

    @property
    def cells(self) -> list[Cell]:
        """
        Get the issue cells with the same names.

        :return: the cells.
        :rtype: list[Cell]
        """
        return [cell for cell in self.excel_prop.table_cells if cell.value == self.issue]

    @property
    def dict_issues(self) -> dict[Cell, int]:
        """
        Get the dictionary of the issue cells and state priorities.

        :return: the cell and state priority mapping.
        :rtype: dict[Cell, int]
        """
        dict_cell_priority: dict[Cell, int] = dict()
        for cell in self.cells:
            state = self.excel_prop.cell_states[cell.coordinate]
            dict_cell_priority[cell] = self.excel_prop.dict_state_priority[state]
        return dict_cell_priority

    def work_item_cells(self) -> list[Cell]:
        """
        Get the work item cells to join.

        :return: the work item cells.
        :rtype: list[Cell]
        """
        work_cells: list[Cell] = []
        for cell in self.cells:
            if cell == self.cell_final():
                continue
            else:
                for item in self.excel_prop.cell_in_range(f"G{cell.row}", f"NR{cell.row}"):
                    if item.value is not None:
                        work_cells.append(item)
                    else:
                        continue
        return work_cells

    def cell_final(self) -> Cell:
        """
        Get the final cell to keep.

        :return: the cell
        :rtype: Cell
        """
        max_priority = max(priority for priority in self.dict_issues.values())
        max_prior_min_row = min(cell.row for cell, priority in self.dict_issues.items() if priority == max_priority)
        for cell in self.cells:
            if cell.row == max_prior_min_row():
                return cell

    def join_cells(self):
        """Join the work items to the final issue."""
        for work_item_cell in self.work_item_cells():
            self.excel_prop.replace_cell(
                from_=work_item_cell, to_=f"{work_item_cell.column_letter}{self.cell_final().row}")
        for issue_cell in self.cells:
            if issue_cell != self.cell_final():
                row = issue_cell.row
                for column in ("B", "C", "D", "E"):
                    self.excel_prop.replace_cell(from_=f"{column}{row}")
