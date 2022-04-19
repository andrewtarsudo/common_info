import calendar
import datetime
import os
import pathlib
import re
from copy import copy
from decimal import Decimal
from typing import Optional, Any, Union
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_THICK
from openpyxl.styles.colors import Color, BLACK, WHITE
from openpyxl.styles.fills import PatternFill, FILL_SOLID
from openpyxl.styles.fonts import Font
from openpyxl.styles.numbers import FORMAT_TEXT, FORMAT_NUMBER_00, FORMAT_DATE_XLSX14, FORMAT_GENERAL
from openpyxl.styles.protection import Protection
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter, coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
import platform


class ConstExcel:
    dict_headers: dict[str, int] = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3, 'Легенда': 4}
    dict_headers_short: dict[str, int] = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3}
    dict_attributes: dict[str, str] = {"parent": "B", "name": "C", "summary": "D", "deadline": "E", "formula": "NS"}


class ConstYT:
    headers_yt = {
        'Authorization': 'Bearer perm:dGFyYXNvdi1h.NjEtMTQw.1udDlV6zaAitHIgvw2eNQvF1sZ9JTZ',
        'Accept': 'application/json',
        'Content-Type': 'application/json',
    }

    os_name = platform.system()

    start_date_default: datetime.date = datetime.date.today() - datetime.timedelta(days=7)

    @classmethod
    def get_user_name(cls):
        if ConstYT.os_name == "Windows":
            param = "USERNAME"
        else:
            param = "USER"

        return os.environ.get(f"{param}")

    @staticmethod
    def convert_date_iso(input_date: str) -> Optional[datetime.date]:
        """
        Converts different date formats to the ISO standard.\n
        :param input_date: the date to convert, str
        :return: the modified date of the str type.
        """
        date_conversion_rules = (
            (re.compile(r'(\d{4}).(\d{1,2}).(\d{1,2})'), (1, 2, 3)),
            (re.compile(r'(\d{1,2}).(\d{1,2}).(\d{4})'), (3, 2, 1))
        )
        print(f'input_date = {input_date}')
        # if the date format is not specified
        if not any(re.match(pattern, input_date) for pattern, match in date_conversion_rules):
            return None

        for pattern, group_match in date_conversion_rules:
            match = re.match(pattern, input_date)
            # find the pattern to convert
            if match:
                i_1, i_2, i_3 = group_match
                date_iso = f'{match.group(i_1)}-{match.group(i_2)}-{match.group(i_3)}'
                return datetime.date.fromisoformat(date_string=date_iso)

    @staticmethod
    def convert_spent_time(spent_time: int) -> Union[int, Decimal]:
        """
        Converts the spent time in minutes to hours.

        :param spent_time: the spent time in minutes, int
        :return: the converted spent time of the Union[int, Decimal] type.
        """
        if spent_time % 60 == 0:
            return spent_time // 60
        else:
            return Decimal(spent_time / 60).normalize()


class ConstStyleCell:
    """Defines the class that contains constant values and styles."""
    # for set_fill_color:
    # type_fill: str, fill_color: str, tint: float, theme: int
    dict_states_colors: dict[str, tuple[str, str, float, int]] = {
        "weekend": ('tmp', 'FFFD5635', None, None), "deadline": ('tmp', 'FFFF0D0D', None, None),
        "done": ('tint', None, 0.0, 9), "active": ('tint', None, 0.3999755851924192, 9),
        "test": ('tmp', 'FF117D68', None, None), "going_start": ('tmp', 'FF00B0F0', None, None),
        "paused": ('tint', None, -0.499984740745262, 0), "verified_closed": ('tint', None, 0.249977111117893, 1),
        "going_finish": ('tint', None, -0.249977111117893, 4,), "sick": ('tint', None, -0.249977111117893, 7),
        "vacation": ('tmp', 'FFFFFF00', None, None)
    }
    # Cell.alignment
    TMP_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrap_text=True)
    CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ROTATE_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True, text_rotation=90)
    # Cell.border
    TMP_BORDER = Border(outline=False, left=Side(), right=Side(), top=Side(), bottom=Side())
    THIN_BORDER = Border(
        outline=True, left=Side(style=BORDER_THIN), right=Side(style=BORDER_THIN), top=Side(style=BORDER_THIN),
        bottom=Side(style=BORDER_THIN))
    THICK_BORDER = Border(
        outline=True, left=Side(style=BORDER_THICK), right=Side(style=BORDER_THICK), top=Side(style=BORDER_THICK),
        bottom=Side(style=BORDER_THICK))
    TOP_BOTTOM_BORDER = Border(
        outline=True, left=Side(style=BORDER_THIN), right=Side(style=BORDER_THIN), top=Side(style=BORDER_THICK),
        bottom=Side(style=BORDER_THICK))
    # Cell.fill
    TMP_FILL = PatternFill(
        fill_type=FILL_SOLID, fgColor=Color(rgb=WHITE, type='rgb'), bgColor=Color(rgb=WHITE, type='rgb'))
    INDEXED_FILL = PatternFill(
        fill_type=FILL_SOLID, fgColor=Color(rgb=WHITE, type='rgb'), bgColor=Color(indexed=64, type='indexed'))
    TINT_FILL = PatternFill(
        fill_type=FILL_SOLID, fgColor=Color(theme=0, tint=0.0, type='theme'), bgColor=Color(indexed=64, type='indexed'))
    # Cell.font
    TMP_FONT = Font(name='Calibri', charset=204, family=2, color=Color(rgb=BLACK, type='rgb'), size=11)
    THEME_FONT = Font(name='Calibri', charset=204, family=2, color=Color(theme=1, type='theme'), size=11)
    # Cell.protection
    TMP_PROTECTION = Protection(locked=False, hidden=False)

    @staticmethod
    def set_number_format(format_type: str = None) -> str:
        """
        Specifies the cell number format.\n
        text: FORMAT_TEXT (@)\n
        number: FORMAT_NUMBER_00 (0,00)\n
        date: FORMAT_DATE_XLSX14 (%d-%m-%Y)\n
        default: FORMAT_GENERAL (General)\n
        :param format_type: the cell value type, str
        :return: the number format of the str type.
        """
        if format_type is None:
            return FORMAT_GENERAL

        if format_type == 'text':
            return FORMAT_TEXT
        elif format_type == 'number':
            return FORMAT_NUMBER_00
        elif format_type == 'date':
            return FORMAT_DATE_XLSX14
        else:
            return FORMAT_GENERAL

    def set_alignment(self, type_align: str = None) -> Alignment:
        """
        Specifies the cell alignment.\n
        center: CENTER_ALIGNMENT (hor = centered, vert = centered)\n
        rotate: ROTATE_ALIGNMENT (hor = centered, vert = centered, text_rot = 90)\n
        default: TMP_ALIGNMENT (hor = left, vert = centered)\n
        :param type_align: the alignment type, str
        :return: the cell alignment of the Alignment() type.
        """
        if type_align is None:
            return self.TMP_ALIGNMENT

        if type_align.lower().strip() == 'center':
            return self.CENTER_ALIGNMENT
        elif type_align.lower().strip() == 'rotate':
            return self.ROTATE_ALIGNMENT
        else:
            return self.TMP_ALIGNMENT

    def set_fill_color(
            self, type_fill: str = None, fill_color: str = None, tint: float = None, theme: int = None) -> PatternFill:
        """
        Specifies the cell fill.\n
        tmp: TMP_FILL = (type_fill='rgb', fill_color=Color(), tint=None, theme=None)\n
        tint: TINT_FILL = (type_fill='theme', fill_color=None, tint=tint, theme=theme)\n
        indexed: INDEXED_FILL = (type_fill='indexed', fill_color=Color(), tint=None, theme=theme)\n
        :param type_fill: the fill type, str
        :param fill_color: the RGB cell fill color, str
        :param tint: the tint value, float
        :param theme: the cell theme to apply, int
        :return: the cell fill of the PatternFill type.
        """
        if all(attribute is None for attribute in (type_fill, fill_color, tint, theme)):
            return self.TMP_FILL

        if type_fill.lower().strip() == 'tmp':
            res_fill: PatternFill = copy(self.TMP_FILL)
            if fill_color is not None:
                res_fill.fgColor.rgb = fill_color

        elif type_fill.lower().strip() == 'tint':
            res_fill: PatternFill = copy(self.TINT_FILL)
            res_fill.fill_type = 'theme'

            if tint is not None:
                res_fill.fgColor.tint = tint
            if theme is not None:
                res_fill.fgColor.theme = theme

        elif type_fill.lower().strip() == 'indexed':
            res_fill: PatternFill = copy(self.INDEXED_FILL)
            res_fill.fill_type = 'indexed'

            if fill_color is not None:
                res_fill.fgColor.rgb = fill_color

        else:
            res_fill: PatternFill = copy(self.TMP_FILL)

        return res_fill

    def set_border(self, type_border: str = None) -> Border:
        """
        Specifies the cell borders.\n
        thin: THIN_BORDER (all = THIN)\n
        thick: THICK_BORDER (all = THICK)\n
        top_bottom: TOP_BOTTOM_BORDER (left + right = THIN, top + bottom = THICK)\n
        :param type_border: the border type, str
        :return: the cell border of the Border() type.
        """
        if type_border is None:
            return self.TMP_BORDER

        if type_border.lower().strip() == 'thin':
            return self.THIN_BORDER
        elif type_border.lower().strip() == 'thick':
            return self.THICK_BORDER
        elif type_border.lower().strip() == 'top_bottom':
            return self.TOP_BOTTOM_BORDER
        else:
            return self.TMP_BORDER

    def set_font(self, type_font: str = None) -> Font:
        """
        Specifies the cell font.\n
        theme: THEME_FONT (Color(type="theme", theme=1))\n
        tmp: TMP_FONT (Color(type="rgb", rgb=BLACK))\n
        :param type_font: the font type, str
        :return: the cell font of the Font() type.
        """
        if type_font is None:
            return self.TMP_FONT

        if type_font.lower().strip() == 'tmp':
            return self.TMP_FONT
        elif type_font.lower().strip() == 'theme':
            return self.THEME_FONT
        else:
            return self.TMP_FONT

    @property
    def set_protection(self) -> Protection:
        """
        Specifies the cell protection.\n
        TMP_PROTECTION(locked=False, hidden=False)\n
        :return: the Protection() type.
        """
        return self.TMP_PROTECTION

    @property
    def month_title_styles(self) -> list[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]]:
        """Specifies the styles of the month titles."""
        month_title_styles: list[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]] = []

        for name, month_params in ConstDefault.dict_months_colors.items():
            type_fill, fill_color, tint, theme = month_params
            # specify cell parameters
            number_format = self.set_number_format(format_type="text")
            alignment = self.set_alignment(type_align='rotate')
            border = self.set_border(type_border='top_bottom')
            fill = self.set_fill_color(type_fill=type_fill, fill_color=fill_color, tint=tint, theme=theme)
            font = self.set_font(type_font='theme')
            protection = self.set_protection

            month_title_styles.append((name, number_format, alignment, border, fill, font, protection))

        return month_title_styles

    @property
    def month_header_styles(self) -> list[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]]:
        """Specifies the styles of the month headers."""
        month_header_styles: list[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]] = []

        for index, month_title_style in enumerate(self.month_title_styles):
            # specify cell parameters
            name, number_format, alignment, border, fill, font, protection = copy(self.month_title_styles[index])
            # the only difference is rotation
            alignment = self.set_alignment(type_align='center')
            month_header_styles.append((name, number_format, alignment, border, fill, font, protection))

        return month_header_styles

    @property
    def state_styles(self) -> list[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]]:
        """Specifies the styles of the states."""
        state_styles: list[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]] = []

        for name, item in self.dict_states_colors.items():
            type_fill, fill_color, tint, theme = item
            # specify cell parameters
            number_format = self.set_number_format(format_type="text")
            alignment = self.set_alignment(type_align='center')
            border = self.set_border(type_border='thick')
            fill = self.set_fill_color(type_fill=type_fill, fill_color=fill_color, tint=tint, theme=theme)
            font = self.set_font(type_font='theme')
            protection = self.set_protection
            state_styles.append((name, number_format, alignment, border, fill, font, protection))

        return state_styles

    @property
    def header_style(self) -> tuple[str, str, Alignment, Border, PatternFill, Font, Protection]:
        """Specifies the style of the headers."""
        # set default values for headers
        name: str = 'headers'
        theme: int = 3
        tint: float = 0.5999938962981048
        # specify cell parameters
        number_format = self.set_number_format(format_type="text")
        alignment = self.set_alignment(type_align='center')
        border = self.set_border(type_border='top_bottom')
        fill = self.set_fill_color(type_fill='tint', tint=tint, theme=theme)
        font = self.set_font(type_font='theme')
        protection = self.set_protection

        return name, number_format, alignment, border, fill, font, protection

    @property
    def title_style(self) -> tuple[str, str, Alignment, Border, PatternFill, Font, Protection]:
        """Specifies the style of the title."""
        # set default values for titles
        name: str = 'title'
        theme: int = 4
        tint: float = -0.249977111117893
        # specify cell parameters
        number_format = self.set_number_format(format_type="text")
        alignment = self.set_alignment(type_align='center')
        border = self.set_border(type_border='top_bottom')
        fill = self.set_fill_color(type_fill='tint', tint=tint, theme=theme)
        font = self.set_font(type_font='theme')
        protection = self.set_protection

        return name, number_format, alignment, border, fill, font, protection

    @property
    def date_style(self):
        """Specifies the style of the dates."""
        # set default values for dates
        name: str = 'date'
        # specify cell parameters
        number_format = self.set_number_format(format_type="date")
        alignment = self.set_alignment(type_align='center')
        border = self.set_border(type_border='top_bottom')
        fill = self.set_fill_color(type_fill='rgb', fill_color=WHITE, tint=None, theme=None)
        font = self.set_font(type_font='tmp')
        protection = self.set_protection

        return name, number_format, alignment, border, fill, font, protection

    @property
    def basic_style(self):
        """Specifies the basic style."""
        # set default values for dates
        name: str = 'basic'
        # specify cell parameters
        number_format = self.set_number_format(format_type="default")
        alignment = self.set_alignment(type_align=None)
        border = self.set_border(type_border=None)
        fill = self.set_fill_color(type_fill=None, fill_color=None, tint=None, theme=None)
        font = self.set_font(type_font=None)
        protection = self.set_protection

        return name, number_format, alignment, border, fill, font, protection

    @property
    def list_all_styles(self):
        """Specifies the list of all styles."""
        all_styles = [*self.all_named_styles, *self.all_cell_styles]
        return *all_styles,

    @property
    def all_style_names(self):
        """Specifies the list of all style names."""
        return [name for style in self.list_all_styles for name, _, _, _, _, _, _ in style]

    @property
    def all_named_styles(self) -> tuple[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]]:
        """Specifies the tuple of all named styles."""
        named_styles = [*self.state_styles, self.header_style, self.basic_style]
        return *named_styles,

    @property
    def named_styles_name(self) -> list[str]:
        """Specifies the names of MamedStyle."""
        return [name for named_style in self.all_named_styles for name, _, _, _, _, _, _ in named_style]

    @property
    def all_cell_styles(self) -> tuple[tuple[str, str, Alignment, Border, PatternFill, Font, Protection]]:
        """Specifies the tuple of all not named styles."""
        cell_styles = [*self.month_title_styles, *self.month_header_styles, self.title_style, self.date_style]
        return *cell_styles,

    @property
    def cell_styles_name(self) -> list[str]:
        """Specifies the names of StyleCell."""
        return [name for cell_style in self.all_cell_styles for name, _, _, _, _, _, _ in cell_style]


class ConstDefault:
    # style_name: (rus_name, cell_style_coordinate, cell_legend_coordinate
    dict_legend = {
        'weekend': ('выходные', 'B21', 'C21'), 'deadline': ('дедлайн', 'B22', 'C22'), 'done': ('done', 'B23', 'C23'),
        'active': ('active', 'B24', 'C24'), 'test': ('test', 'B25', 'C25'),
        'going_start': ('планирую начать', 'B26', 'C26'), 'paused': ('paused', 'B27', 'C27'),
        'verified_closed': ('verified, closed', 'B28', 'C28'), 'going_finish': ('планирую завершить', 'B29', 'C29'),
        'sick': ('больничный', 'B30', 'C30'), 'vacation': ('отпуск', 'B31', 'C31')
    }
    # title_name: (rus_name, cell_title_coordinate)
    dict_titles = {
        'parent': ('РОД. ЗАДАЧА', 'B2'), 'name': ('ЗАДАЧА', 'C2'), 'summary': ('ОПИСАНИЕ', 'D2'),
        'deadline': ('DEADLINE', 'E2'), 'sum': ('Σ', 'NS2'), 'commentary': ('ПРИМЕЧАНИЕ', 'NT2')
    }
    # month_name: (rus_name, cell_title_coordinate, num_days)
    dict_month_ranges: dict[str, tuple[str, str, int]] = {
        'january': ('ЯНВАРЬ', 'F1', 31), 'february': ('ФЕВРАЛЬ', 'AL1', 28), 'march': ('МАРТ', 'BO1', 31),
        'april': ('АПРЕЛЬ', 'CU1', 30), 'may': ('МАЙ', 'DZ1', 31), 'june': ('ИЮНЬ', 'FF1', 30),
        'july': ('ИЮЛЬ', 'GK1', 31), 'august': ('АВГУСТ', 'HQ1', 31), 'september': ('СЕНТЯБРЬ', 'IW1', 30),
        'october': ('ОКТЯБРЬ', 'KB1', 31), 'november': ('НОЯБРЬ', 'LH1', 30), 'december': ('ДЕКАБРЬ', 'MM1', 31)
    }

    cells_headers: tuple[tuple[int, str]] = ((3, 'Active'), (7, 'New/Paused'), (11, 'Active/Done'), (15, 'Verified'))
    # holidays
    yy: int = datetime.date.today().year
    holidays = (
        datetime.date(yy, 1, 1), datetime.date(yy, 1, 2), datetime.date(yy, 1, 3), datetime.date(yy, 1, 4),
        datetime.date(yy, 1, 5), datetime.date(yy, 1, 6), datetime.date(yy, 1, 7), datetime.date(yy, 2, 23),
        datetime.date(yy, 3, 7), datetime.date(yy, 3, 8), datetime.date(yy, 5, 1), datetime.date(yy, 5, 2),
        datetime.date(yy, 5, 3), datetime.date(yy, 5, 8), datetime.date(yy, 5, 9), datetime.date(yy, 5, 10),
        datetime.date(yy, 6, 12), datetime.date(yy, 6, 13), datetime.date(yy, 11, 4), datetime.date(yy, 12, 31)
    )
    # for set_fill_color:
    # type_fill: str, fill_color: str, tint: float = None, theme: int = None
    dict_months_colors = {
        'january': ('tmp', 'FF0297CC', None, None), 'february': ('tmp', 'FF0BA6B6', None, None),
        'march': ('tmp', 'FF3AAA66', None, None), 'april': ('tmp', 'FF8BBD36', None, None),
        'may': ('tmp', 'FFD0CA04', None, None), 'june': ('tmp', 'FFF9AD01', None, None),
        'july': ('tmp', 'FFF08002', None, None), 'august': ('tmp', 'FFE94442', None, None),
        'september': ('tmp', 'FFCF687D', None, None), 'october': ('tmp', 'FF98668B', None, None),
        'november': ('tmp', 'FF697FB9', None, None), 'december': ('tmp', 'FF0078A9', None, None)
    }
    # month title StyleCell values
    # ['january', 'february', 'march', 'april', 'may', 'june', 'july',
    # 'august', 'september', 'october', 'november', 'december']
    dict_month_title_styles: dict[str, tuple] = dict()
    # month header StyleCell values
    dict_month_header_styles: dict[str, tuple] = dict()
    # StyleCell values
    dict_style_cells: dict[str, tuple] = dict()


class Const:
    @staticmethod
    def cell_column(coord: str) -> str:
        """
        Gets the cell column.\n
        :param coord: the cell coordinate, str
        :return: the column of the str type.
        """
        return coordinate_from_string(coord)[0]

    @staticmethod
    def cell_row(coord: str) -> int:
        """
        Gets the cell row.\n
        :param coord: the cell coordinate, str
        :return: the row of the int type.
        """
        return coordinate_from_string(coord)[1]

    @staticmethod
    def add_column(coord: str, add_value: int):
        """
        Finds the cell in the row that is N columns to the left/right.\n
        :param coord: the base cell coordinate, str
        :param add_value: the number of columns with the sign, int
        :return: the new cell coordinate of the str value.
        """
        column, row = coordinate_from_string(coord)
        col_idx = column_index_from_string(column)
        if add_value < 0 and col_idx + add_value < 1:
            print('The incorrect work of the add_column method.')
            return coord
        add_col_idx = col_idx + add_value
        add_column = get_column_letter(add_col_idx).upper()
        return f'{add_column}{row}'

    @staticmethod
    def cell_in_range(start_coord: str, end_coord: str):
        """
        Converts the cell range to the cell generator in the range.\n
        :param start_coord: the start cell coordinate, str
        :param end_coord: the end cell coordinate, str
        :return: the generator of cells in the cell range.
        """
        if Const.check_coord(coord=start_coord) and Const.check_coord(coord=end_coord):
            min_row, max_row, min_col, max_col = Const.range_coord(start_coord=start_coord, end_coord=end_coord)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    yield f'{get_column_letter(col)}{row}'

    @staticmethod
    def check_coord(coord: str) -> bool:
        """
        Verifies the cell coordinate.\n
        :param coord: the cell coordinate, str
        :return: the verification flag of the bool type.
        """
        flag = False
        try:
            coordinate_to_tuple(coord)
        except TypeError as e:
            print(f'TypeError occurred. Error in line {e.__traceback__.tb_lineno}. Incorrect value type.\n')
            print(f'coordinate = {coord}')
        except ValueError as e:
            print(f'ValueError occurred. Error in line {e.__traceback__.tb_lineno}. Incorrect value.\n')
            print(f'coordinate = {coord}')
        except OSError as e:
            print(f'OSError occurred. Error {e.errno} in line {e.__traceback__.tb_lineno}.\n')
            print(f'coordinate = {coord}')
        else:
            flag = True
        finally:
            return flag

    @staticmethod
    def range_coord(start_coord: str, end_coord: str):
        """
        Convert the range string with the start and end coordinates to the tuple:\n
        (min_row, min_col, max_row, max_col)\n
        :param start_coord: the start cell coordinate, str
        :param end_coord: the end cell coordinate, str
        :return: the values for iteration of the tuple[int, int, int, int] type.
        """
        # convert start coordinate
        start_row, start_column = coordinate_to_tuple(start_coord)
        # start_col_idx = column_index_from_string(start_column)
        # convert end coordinate
        end_row, end_column = coordinate_to_tuple(end_coord)
        # end_col_idx = column_index_from_string(end_column)
        # find min and max values
        min_row = min(start_row, end_row)
        max_row = max(start_row, end_row)
        # min_col = min(start_col_idx, end_col_idx)
        # max_col = max(start_col_idx, end_col_idx)
        min_col = min(start_column, end_column)
        max_col = max(start_column, end_column)

        return min_row, max_row, min_col, max_col

    @staticmethod
    def convert_datetime_date(value: Any) -> Optional[datetime.date]:
        """
        Converts the datetime or None to the date.\n
        :param value: the value to convert, Any
        :return: the converted value or None of the date type.
        """
        if isinstance(value, datetime.date):
            return value
        elif isinstance(value, datetime.datetime):
            return value.date()
        else:
            return None

    @classmethod
    def set_cells_weekend(cls, ws: Worksheet) -> tuple[str]:
        """
        Specifies the coordinates to set the weekend style.
        :param ws: the current worksheet, Worksheet
        :return: the number of
        """
        holidays_coord = set()
        for coord in cls.cell_in_range(start_coord='G1', end_coord='NR1'):
            cell_value: Optional[datetime.date] = cls.convert_datetime_date(ws[f'{coord}'].value)
            year, month, day = (cell_value.year, cell_value.month, cell_value.day)
            if cell_value is not None:
                if calendar.weekday(year=year, month=month, day=day) in (5, 6) or cell_value in ConstDefault.holidays:
                    holidays_coord.add(coord)
        return *holidays_coord,

    @staticmethod
    def check_terminate_script(prompt: str) -> str:
        """
        Check the input to terminate the program.\n
        :param prompt: the text to display, str
        :return: either exit() or modified command of the str type.
        """
        terminate_commands = ("__exit__", "'__exit__'", '"__exit__"')
        user_input = input(prompt)
        # delete trailing zeros and lower case
        user_command = user_input.lower().strip()

        if user_command in terminate_commands:
            print('Работа прервана. Программа закрывается.')
            exit()
        else:
            return user_command

    @staticmethod
    def set_path_file(path: Union[str, pathlib.Path]) -> bool:
        """
        Defines the path of the file to save the results.\n
        :return: the name of the file and the absolute path of the tuple[str, Path] type.
        """
        try:
            with open(path, 'w+') as file:
                file.read()
        except FileNotFoundError:
            print('Такого пути не существует.')
            return False
        except RuntimeError:
            print('Указанный путь некорректен.')
            return False
        except OSError as e:
            print(f'Произошла ошибка {e.__class__.__name__}. Сообщите об ошибке. Спасибо.')
            return False
        else:
            return True

    @staticmethod
    def merge_cells(ws: Worksheet, row: int):
        ws.merge_cells(f'B{row}:E{row}')

    @staticmethod
    def unmerge_cells(ws: Worksheet, row: int):
        ws.unmerge_cells(f'B{row}:E{row}')

    @classmethod
    def month_cell_periods(cls):
        """Merges cells for the month title."""
        start_cells = [Const.add_column(month_coord, 1) for month_params in ConstDefault.dict_month_ranges.values()
                       for _, month_coord, _ in month_params]
        end_cells = [Const.add_column(month_coord, num_days) for month_params in ConstDefault.dict_month_ranges.values()
                     for _, month_coord, num_days in month_params]
        return [f"{start_cell}:{end_cell}" for start_cell, end_cell in zip(start_cells, end_cells)]

    @classmethod
    def month_cell_periods_start(cls):
        return [Const.add_column(month_coord, 1) for month_params in ConstDefault.dict_month_ranges.values()
                for _, month_coord, _ in month_params]


def main():
    print(ConstDefault.dict_legend.keys())


if __name__ == '__main__':
    main()
