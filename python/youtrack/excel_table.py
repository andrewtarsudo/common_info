from copy import copy
from typing import Optional, Union, Any
import datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_to_tuple, coordinate_from_string, get_column_letter
from decimal import Decimal
from _style_work_item import _StyleWorkItem, _StyleWorkItemList
from openpyxl.styles.cell_style import CellStyle


class ConstXL:
    dict_excel_prop = dict()
    dict_pyxl_row = dict()
    dict_pyxl_work_item = dict()


def reverse_dict(dict_init: dict):
    dict_reverse = dict()
    for key, value in dict_init.items():
        dict_reverse[value] = key
    return dict_reverse


class ExcelProp:
    dict_headers = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3, 'Легенда': 4}
    dict_headers_short = {'Active': 0, 'New/Paused': 1, 'Done/Test': 2, 'Verified': 3}

    def __init__(self, ws: Worksheet, name: str):
        self.ws = ws
        self.name = name

        ConstXL.dict_excel_prop[self.name] = self

    def __str__(self):
        return f"Worksheet is {self.ws}, name is {self.name}"

    def __repr__(self):
        return f"ExcelProp(ws={self.ws}, name={self.name})"

    def __hash__(self):
        return hash((self.ws, self.name))

    def __eq__(self, other):
        if isinstance(other, ExcelProp):
            return (self.ws == other.ws) and (self.name == other.name)
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, ExcelProp):
            return (self.ws != other.ws) and (self.name != other.name)
        else:
            return NotImplemented

    @property
    def bottom_right(self):
        """"""
        max_row = self.ws.max_row
        max_col = self.ws.max_column
        return self.ws.cell(max_row, max_col)

    @property
    def max_column(self):
        return self.bottom_right.column

    @property
    def max_column_letter(self):
        """"""
        return self.bottom_right.column_letter

    @property
    def max_row(self):
        """"""
        return self.bottom_right.row

    def cell_from_coord(self, coord: str) -> Cell:
        """"""
        return self.ws[f"{coord}"]

    def cell_in_range(self, start_coord: str, end_coord: str):
        """
        Converts the cell range to the cell generator in the range.\n
        :param start_coord: the start cell coordinate, str
        :param end_coord: the end cell coordinate, str
        :return: the generator of cells of the Cell type.
        """
        if check_coord(coord=start_coord) and check_coord(coord=end_coord):
            min_row, max_row, min_col, max_col = range_coord(start_coord=start_coord, end_coord=end_coord)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    coord = f'{get_column_letter(col)}{row}'
                    yield self.ws[coord]

    @property
    def headers(self) -> list[Cell]:
        """"""
        end_coord = f"B{self.bottom_right.row}"
        return [cell for cell in self.cell_in_range("B3", end_coord) if cell.value in ExcelProp.dict_headers]

    @property
    def headers_row(self) -> list[int]:
        """"""
        return [header.row for header in self.headers]

    @staticmethod
    def column_coord(coord: str) -> str:
        """"""
        return coordinate_from_string(coord)[0]

    @staticmethod
    def row_coord(coord: str) -> int:
        """"""
        return coordinate_from_string(coord)[1]

    def check_empty(self, item: Union[int, str, Cell]):
        """
        Checks if the cell is empty. The item may be:\n
        the cell, Cell;\n
        the cell coordinate, str;\n
        the row, int.\n
        :param item: the cell parameter, Union[int, str, Cell]
        :return: the flag if the cell is empty of the bool type.
        """
        if isinstance(item, Cell):
            return True if item.value is None else False
        elif isinstance(item, str):
            return True if self.ws[f"{item}"].value is None else False
        elif isinstance(item, int):
            if self.ws[f"B{item}"] | self.ws[f"C{item}"] | self.ws[f"D{item}"] | self.ws[f"E{item}"] is not None:
                return False
            else:
                return True

    def list_state_item(self, state: str) -> list[int]:
        """

        :param state:
        :return:
        """
        index = self.dict_headers_short[state]
        return [row for row in range(self.headers_row[index], self.headers_row[index + 1]) if not self.check_empty(row)]

    @property
    def list_state_row(self) -> list[list[int]]:
        """"""
        return [self.list_state_item(state) for state, index in self.dict_headers_short.items()]

    @property
    def tasks(self):
        """"""
        return [PyXLRow.get_pyxl_row(self.name, f"C{row}") for list_state in self.list_state_row for row in list_state]

    def get_work_items(self, row: int) -> list[tuple[datetime.date, Union[int, Decimal], str, str]]:
        """
        Gets the work items.\n
        :param row: the table row, int
        :return: the list of work items of the tuple[date, Union[int, Decimal], str, str] type.
        """
        work_items: list[tuple[datetime.date, Union[int, Decimal], str, str]] = []
        work_item_row: tuple[Cell]
        cell: Cell
        for cell in self.cell_in_range(start_coord=f"G{row}", end_coord=f"NR{row}"):
            # get the cell column
            column = cell.column_letter
            if cell.value is None:
                continue
            else:
                # get style
                if cell.has_style:
                    cell_style: str = cell._style.name
                else:
                    cell_style = "basic"
                work_items.append(
                    (self.ws[f'{column}1'].value, cell.value, cell.coordinate, cell_style))
        return work_items


def cell_column(coord: str) -> str:
    """
    Gets the cell column.\n
    :param coord: the cell coordinate, str
    :return: the column of the str type.
    """
    return coordinate_from_string(coord)[0]


def cell_row(coord: str) -> int:
    """
    Gets the cell row.\n
    :param coord: the cell coordinate, str
    :return: the row of the int type.
    """
    return coordinate_from_string(coord)[1]


def check_coord(coord: str) -> bool:
    """
    Verifies the cell coordinate.\n
    :param coord: the cell coordinate, str
    :return: the verification flag of the bool type.
    """
    flag = False
    try:
        coordinate_to_tuple(coord)
    except TypeError as e:
        print(f'TypeError occurred. Error in line {e.__traceback__.tb_lineno}. Incorrect value type.\n')
        print(f'coordinate = {coord}')
    except ValueError as e:
        print(f'ValueError occurred. Error in line {e.__traceback__.tb_lineno}. Incorrect value.\n')
        print(f'coordinate = {coord}')
    except OSError as e:
        print(f'OSError occurred. Error {e.errno} in line {e.__traceback__.tb_lineno}.\n')
        print(f'coordinate = {coord}')
    else:
        flag = True
    finally:
        return flag


def range_coord(start_coord: str, end_coord: str):
    """
    Convert the range string with the start and end coordinates to the tuple:\n
    (min_row, min_col, max_row, max_col)\n
    :param start_coord: the start cell coordinate, str
    :param end_coord: the end cell coordinate, str
    :return: the values for iteration of the tuple[int, int, int, int] type.
    """
    # convert start coordinate
    start_row, start_column = coordinate_to_tuple(start_coord)
    # start_col_idx = column_index_from_string(start_column)
    # convert end coordinate
    end_row, end_column = coordinate_to_tuple(end_coord)
    # end_col_idx = column_index_from_string(end_column)
    # find min and max values
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)
    # min_col = min(start_col_idx, end_col_idx)
    # max_col = max(start_col_idx, end_col_idx)
    min_col = min(start_column, end_column)
    max_col = max(start_column, end_column)
    return min_row, max_row, min_col, max_col


def convert_spent_time(spent_time: Any) -> Union[int, Decimal]:
    """Converts the spent time to the specified format."""
    return spent_time if isinstance(spent_time, int) else Decimal(spent_time).normalize()


def converter_work_items(
        work_items: list[tuple[Any, Any, str, str]]) -> list[tuple[datetime.date, Union[int, Decimal], str, str]]:
    work_items_final: list[tuple[datetime.date, Union[int, Decimal], str, str]] = []
    for work_item in work_items:
        date_init, spent_time_init, coord, style_name = work_item
        date_final: Optional[datetime.date] = convert_datetime_date(date_init)
        spent_time_final = convert_spent_time(spent_time_init)
        work_items_final.append((date_final, spent_time_final, coord, style_name))
    return work_items_final


def convert_datetime_date(value: Any) -> Optional[datetime.date]:
    """
    Converts the datetime or None to the date.\n
    :param value: the value to convert, Any
    :return: the converted value or None of the date type.
    """
    if isinstance(value, datetime.date):
        return value
    elif isinstance(value, datetime.datetime):
        return value.date()
    else:
        return None


class PyXLRow:
    index = 0
    __slots__ = ("excel_prop_name", "issue", "identifier")

    def __init__(self, excel_prop_name: str, issue: Cell):
        self.excel_prop_name = excel_prop_name
        self.issue = issue
        self.identifier = PyXLRow.index

        ConstXL.dict_pyxl_row[self.identifier] = self
        PyXLRow.index += 1

    def __str__(self):
        return f"PyXLRow: excel_prop_name = {self.excel_prop_name}, cell = {self.issue.coordinate}," \
               f"identifier = {self.identifier}"

    def __repr__(self):
        return f"PyXLRow(excel_prop_name={self.excel_prop_name}, issue={self.issue.coordinate})," \
               f"PyXLRow.identifier={self.identifier}"

    def __hash__(self):
        return hash((self.excel_prop_name, self.issue.coordinate))

    def __key(self):
        """Sets the private tuple of attributes to use next."""
        return self.excel_prop_name, self.issue.coordinate

    def __eq__(self, other):
        if isinstance(other, PyXLRow):
            return self.__key() == other.__key()
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, PyXLRow):
            return self.__key() != other.__key()
        else:
            return NotImplemented

    def __lt__(self, other):
        if isinstance(other, PyXLRow):
            return self.issue.row < other.issue.row
        else:
            return NotImplemented

    def __gt__(self, other):
        if isinstance(other, PyXLRow):
            return self.issue.row > other.issue.row
        else:
            return NotImplemented

    def __le__(self, other):
        if isinstance(other, PyXLRow):
            return self.issue.row <= other.issue.row
        else:
            return NotImplemented

    def __ge__(self, other):
        if isinstance(other, PyXLRow):
            return self.issue.row >= other.issue.row
        else:
            return NotImplemented

    @classmethod
    def get_pyxl_row(cls, excel_prop_name: str, coord: str):
        """
        Gets the class instance.\n
        :param excel_prop_name: the ExcelProp instance name, str
        :param coord: the cell coordinate, str
        :return: the instance of the PyXLRow
        """
        excel_prop = ConstXL.dict_excel_prop[excel_prop_name]
        return cls(excel_prop_name, excel_prop.ws[f"{coord}"])

    @property
    def excel_prop(self):
        """Returns the ExcelProp instance."""
        return ConstXL.dict_excel_prop[self.excel_prop_name]

    @property
    def ws(self):
        """Returns the Worksheet instance."""
        return self.excel_prop.ws

    @property
    def coord(self):
        """Returns the cell coordinate."""
        return self.issue.coordinate

    @property
    def row(self):
        """Returns the cell row."""
        return self.issue.row

    @property
    def parent(self):
        """Gets the parent issue name."""
        return self.ws[f"B{self.row}"].value

    @property
    def issue_name(self):
        """Gets the issue name."""
        return self.ws[f"C{self.row}"].value

    @property
    def summary(self):
        """Gets the issue summary."""
        return self.ws[f"D{self.row}"].value

    @property
    def deadline(self):
        """Gets the issue deadline."""
        return self.ws[f"E{self.row}"].value

    @property
    def commentary(self):
        """Gets the issue commentary."""
        return self.ws[f"NS{self.row}"].value

    def __add__(self, other):
        if isinstance(other, int):
            row, column = coordinate_to_tuple(self.coord)
            if column + other < 1:
                raise ValueError("Incorrect operand.")
            else:
                col_idx = column + other
                return self.ws[f"{get_column_letter(col_idx)}{row}"]
        else:
            return NotImplemented

    def __iadd__(self, other):
        if isinstance(other, int):
            return self.__add__(other)
        else:
            return NotImplemented


class PyXLWorkItem:
    index = 0
    attrs = ("number_format", "alignment", "border", "fill", "font", "protection", "data_type")

    __slots__ = ("excel_prop_name", "cell", "identifier")

    def __init__(self, excel_prop_name: str, cell: Cell):
        self.excel_prop_name = excel_prop_name
        self.cell = cell
        self.identifier = PyXLWorkItem.index

        ConstXL.dict_pyxl_work_item[self.identifier] = self
        PyXLWorkItem.index += 1

    def __str__(self):
        return f"excel_prop_name = {self.excel_prop_name}, cell = {self.cell}, identifier = {self.identifier}"

    def __repr__(self):
        return f"PyXLWorkItem(excel_prop_name={self.excel_prop_name}, cell={self.cell}), identifier={self.identifier}"

    def __hash__(self):
        return hash((self.excel_prop_name, self.cell.coordinate))

    def __key(self):
        """Sets the private tuple of attributes to use next."""
        return self.excel_prop_name, self.cell.row, self.cell.column_letter

    def __eq__(self, other):
        if isinstance(other, PyXLWorkItem):
            return self.__key() == other.__key()
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, PyXLWorkItem):
            return self.__key() != other.__key()
        else:
            return NotImplemented

    def __lt__(self, other):
        if isinstance(other, PyXLWorkItem) and self.cell.row == other.cell.row:
            return self.cell.col_idx < other.cell.col_idx
        else:
            return NotImplemented

    def __gt__(self, other):
        if isinstance(other, PyXLWorkItem) and self.cell.row == other.cell.row:
            return self.cell.col_idx > other.cell.col_idx
        else:
            return NotImplemented

    def __le__(self, other):
        if isinstance(other, PyXLWorkItem) and self.cell.row == other.cell.row:
            return self.cell.col_idx <= other.cell.col_idx
        else:
            return NotImplemented

    def __ge__(self, other):
        if isinstance(other, PyXLWorkItem) and self.cell.row == other.cell.row:
            return self.cell.col_idx >= other.cell.col_idx
        else:
            return NotImplemented

    @property
    def excel_prop(self) -> ExcelProp:
        """Gets the ExcelProp instance."""
        return ConstXL.dict_excel_prop[self.excel_prop_name]

    @property
    def ws(self) -> Worksheet:
        """Returns the Worksheet instance."""
        return self.excel_prop.ws

    @property
    def coord(self) -> str:
        """Returns the cell coordinate."""
        return self.cell.coordinate

    @property
    def row(self) -> int:
        """Returns the cell row."""
        return self.cell.row

    @property
    def spent_time(self) -> Union[int, Decimal]:
        """Gets the work item spent time."""
        return convert_spent_time(self.cell.value)

    @property
    def date(self):
        """Gets the work item date."""
        column: str = self.cell.column_letter
        return convert_datetime_date(self.ws[f"{column}1"].value)

    @property
    def _get_cell_style(self) -> Optional[str]:
        """Gets the cell style."""
        if not self.cell.has_style:
            return None
        else:
            cell_style: CellStyle = self.cell.style
            # get the style name
            if cell_style.name in _StyleWorkItemList.style_names:
                return cell_style.name
            else:
                _style_work_item = _StyleWorkItem.get_style_cell(cell_style.name, self.cell)
                return _style_work_item.name

    def set_style(self, style_name: str):
        """
        Sets the cell style.\n
        :param style_name: the style name, str
        :return: None
        """
        if style_name not in _StyleWorkItemList.style_names:
            self.cell._style = copy(_StyleWorkItemList.basic())
        else:
            self.cell._style = copy(_StyleWorkItemList.get_style(style_name))

    @property
    def cell_style_params(self):
        """Gets the cell attributes."""
        return [self.cell.__getattribute__(attr) for attr in PyXLWorkItem.attrs if attr in PyXLWorkItem.attrs]

    def _set_cell_attr(self, attr: str, value):
        """
        Sets the cell attributes.\n
        :param attr: the cell attribute, str
        :param value: the attribute value, Any
        :return: None.
        """
        if attr in PyXLWorkItem.attrs:
            self.cell.__setattr__(attr, value)
        else:
            print(f"Attribute {attr} is not defined.")


class _PyXLMerged:
    __slots__ = ("excel_prop_name", "cell")

    def __init__(self, excel_prop_name: str, cell: Cell):
        self.excel_prop_name = excel_prop_name
        self.cell = cell

    @property
    def excel_prop(self) -> ExcelProp:
        """Gets the ExcelProp instance."""
        return ConstXL.dict_excel_prop[self.excel_prop_name]

    @property
    def __coord(self) -> str:
        """Gets the cell coordinate."""
        return self.cell.coordinate

    @property
    def row(self) -> int:
        """Gets the cell row."""
        return self.cell.row

    @property
    def ws(self) -> Worksheet:
        """Gets the Worksheet instance."""
        return self.excel_prop.ws

    @property
    def pyxl_row(self) -> int:
        """Gets the PyXLRow instance."""
        pyxl_row: PyXLRow
        for pyxl_row_id, pyxl_row in ConstXL.dict_pyxl_row.items():
            if self.cell == pyxl_row.issue:
                return pyxl_row_id

    @property
    def work_items(self) -> list[int]:
        """Gets the PyXLWorkItem instance."""
        return [work_item_id for work_item_id, work_item in ConstXL.dict_pyxl_work_item.items()
                if work_item.row == self.row]

    def __hash__(self):
        return hash(self.__coord)

    def __iter__(self):
        for work_item in self.work_items:
            yield self.__getitem_id(work_item)

    def __getitem_id(self, identifier: int) -> Optional[PyXLWorkItem]:
        """
        Gets the PyXLWorkItem instance.\n
        :param identifier: the instance identifier, int
        :return: the instance of the PyXLWorkItem type.
        """
        if identifier in self.work_items:
            return ConstXL.dict_pyxl_work_item[identifier]
        else:
            return None

    def __eq__(self, other):
        if isinstance(other, _PyXLMerged):
            return self.__coord == other.__coord
        else:
            return NotImplemented

    def __ne__(self, other):
        if isinstance(other, _PyXLMerged):
            return self.__coord != other.__coord
        else:
            return NotImplemented

    def __lt__(self, other):
        if isinstance(other, _PyXLMerged):
            return self.row < other.row
        else:
            return NotImplemented

    def __gt__(self, other):
        if isinstance(other, _PyXLMerged):
            return self.row > other.row
        else:
            return NotImplemented

    def __le__(self, other):
        if isinstance(other, _PyXLMerged):
            return self.row <= other.row
        else:
            return NotImplemented

    def __ge__(self, other):
        if isinstance(other, _PyXLMerged):
            return self.row >= other.row
        else:
            return NotImplemented

    def __getattribute__(self, item):
        if item in self.__slots__:
            return object.__getattribute__(self, item)
        return None

    def __setattr__(self, key, value):
        if key in self.__slots__:
            object.__setattr__(self, key, value)
        else:
            raise AttributeError(f"Incorrect attribute. Only {self.__slots__} are allowed.")

    def __getitem__(self, item):
        if item in self.work_items:
            return self.work_items[item]
        return None

    def __setitem__(self, key, value):
        self.work_items[key] = value

    def __contains__(self, item):
        if isinstance(item, PyXLRow):
            return self.pyxl_row == item.identifier
        elif isinstance(item, PyXLWorkItem):
            return item.identifier in self.work_items
        else:
            return NotImplemented

    @property
    def __get_pyxl_row(self) -> PyXLRow:
        """Gets the PyXLRow instance."""
        return ConstXL.dict_pyxl_row[self.pyxl_row]

    @property
    def parent(self) -> Optional[str]:
        """Gets the parent issue name."""
        return self.__get_pyxl_row.parent

    @property
    def issue_name(self) -> Optional[str]:
        """Gets the issue name."""
        return self.__get_pyxl_row.issue_name

    @property
    def summary(self) -> Optional[str]:
        """Gets the issue summary."""
        return self.__get_pyxl_row.summary

    @property
    def deadline(self) -> Optional[datetime.date]:
        """Gets the issue deadline."""
        return self.__get_pyxl_row.deadline

    @property
    def commentary(self) -> Optional[str]:
        """Gets the issue commentary."""
        return self.__get_pyxl_row.commentary

    @property
    def item_params(self):
        """Gets the work item parameters to compare."""
        return zip(self.date, self.spent_time)

    @property
    def full_params(self):
        """Gets the work item parameters."""
        return zip(self.date, self.spent_time, self.coord, self.style)

    @property
    def date(self) -> list[datetime.date]:
        """Gets the work item dates."""
        return [self.__getitem_id(identifier).date for identifier in ConstXL.dict_pyxl_work_item]

    @property
    def spent_time(self) -> list[Union[int, Decimal]]:
        """Gets the work item spent time values."""
        return [self.__getitem_id(identifier).spent_time for identifier in ConstXL.dict_pyxl_work_item]

    @property
    def coord(self):
        """Gets the work item coordinates."""
        return [self.__getitem_id(identifier).coord for identifier in ConstXL.dict_pyxl_work_item]

    @property
    def style(self):
        """Gets the work item styles."""
        return [self.__getitem_id(identifier).cell.style for identifier in ConstXL.dict_pyxl_work_item]

    def __get_item_attr(self, attr: str):
        """Gets the work item attribute."""
        list_attr = ("date", "spent_time", "coord", "style")
        if attr not in list_attr:
            return None
        else:
            if attr == "date":
                return self.date
            if attr == "spent_time":
                return self.spent_time
            if attr == "coord":
                return self.coord
            if attr == "style":
                return self.style


def main():
    pass


if __name__ == "__main__":
    main()
